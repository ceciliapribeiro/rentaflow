"""
smart_aporte.py
Smart Aporte Pro — Gestão Ativa de Carteira com Sugestão de Compra e Venda.

Lógica de compra:
  - Prioriza ativos com maior défice de alocação ponderado por DY
  - Score = défice * (1 + DY/100) * fator_pvp
  - Não compra ativo que disparou critério de venda

Lógica de venda (melhores práticas de renda passiva B3):
  1. SOBREALINHAMENTO    — peso real > alvo * (1 + LIMIAR_SOBREPESO)
  2. DY MUITO BAIXO      — DY < DY_MINIMO e há alternativas melhores
  3. P/VP ELEVADO (FII)  — P/VP > PVP_MAXIMO_FII (paga caro demais pelo patrimônio)
  4. PERDA EXCESSIVA     — preço atual < PM * (1 - STOP_PERDA) → stop loss
  5. GANHO EXPRESSIVO    — preço atual > PM * (1 + TAKE_PROFIT) → realizacao parcial

Colunas esperadas:
  Aba CARTEIRA  : col L=ativo, col M=qtde, col P=valor_atual
  Aba Dados B3  : col A=ticker, col D=preço, col G=DY, col H=P/VP
"""

import openpyxl
import math
import os
import sys
import threading
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config_loader import carregar_config

_cfg          = carregar_config()
ARQUIVO_EXCEL = _cfg['arquivo_excel']

# ══════════════════════════════════════════════════════════════════════
# CONFIGURAÇÃO DE COLUNAS DA PLANILHA
# ══════════════════════════════════════════════════════════════════════
ABA_CARTEIRA  = 'CARTEIRA'
ABA_DADOS     = 'Dados B3'

COL_ATIVO_CART   = 12   # L — ticker na aba CARTEIRA
COL_QTDE_CART    = 13   # M — quantidade atual
COL_PM_CART      = 14   # N — preço médio (coluna nova; ajuste se necessário)
COL_VALOR_CART   = 16   # P — valor atual
COL_TICKER_DADOS = 1    # A — ticker na aba Dados B3
COL_PRECO_DADOS  = 4    # D — preço de fechamento
COL_DY_DADOS     = 7    # G — Dividend Yield (%)
COL_PVP_DADOS    = 8    # H — P/VP (Preço sobre Valor Patrimonial)

LINHA_INICIO     = 3    # primeira linha de dados em CARTEIRA

# ══════════════════════════════════════════════════════════════════════
# PARÂMETROS DE DECISÃO — ajuste conforme perfil de risco
# ══════════════════════════════════════════════════════════════════════
LIMIAR_SOBREPESO  = 0.30   # 30 % acima do alvo dispara sugestão de venda
DY_MINIMO         = 4.0    # DY < 4 % é baixo para renda passiva
PVP_MAXIMO_FII    = 1.20   # P/VP > 1,20 = FII caro (>20 % sobre patrimônio)
STOP_PERDA        = 0.25   # preço < PM * (1 - 25 %) = stop loss
TAKE_PROFIT       = 0.40   # preço > PM * (1 + 40 %) = realização parcial
FRACAO_VENDA      = 0.33   # vende até 1/3 da posição por sugestão
MIN_COTAS_VENDA   = 1      # mínimo de cotas para aparecer na boleta de venda

# Paleta de cores
COR_VERDE      = '#1a6b45'
COR_VERDE_ESC  = '#145535'
COR_AZUL       = '#1e3a5f'
COR_VERM       = '#991b1b'
COR_VERM_ESC   = '#7f1d1d'
COR_AMARELO    = '#b45309'
COR_CINZA_BG   = '#f4f4f1'


# ══════════════════════════════════════════════════════════════════════
# UTILITÁRIOS
# ══════════════════════════════════════════════════════════════════════

def limpar_valor(valor):
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    s = str(valor).strip().replace('R$', '').replace('%', '').strip()
    if s.count('.') > 1:
        partes = s.split('.')
        s = ''.join(partes[:-1]) + '.' + partes[-1]
    s = s.replace(',', '.')
    try:
        return float(s)
    except Exception:
        return 0.0


def e_fii(ticker: str) -> bool:
    """Heurística simples: tickers de FII têm 4 letras + 11 (ex: MXRF11)."""
    t = ticker.upper().strip()
    return len(t) == 6 and t[4:].startswith('11')


# ══════════════════════════════════════════════════════════════════════
# LEITURA DA PLANILHA
# ══════════════════════════════════════════════════════════════════════

def carregar_dados_b3():
    base_dados = {}
    ultima_att = None
    if not os.path.isfile(ARQUIVO_EXCEL):
        raise RuntimeError(
            f"Planilha não encontrada:\n{ARQUIVO_EXCEL}\n\n"
            "Abra as Configurações e verifique o caminho do arquivo."
        )
    ultima_att = datetime.fromtimestamp(os.path.getmtime(ARQUIVO_EXCEL))
    wb    = openpyxl.load_workbook(ARQUIVO_EXCEL, data_only=True)
    sheet = wb[ABA_DADOS]
    for row in range(2, sheet.max_row + 1):
        t = sheet.cell(row=row, column=COL_TICKER_DADOS).value
        if t:
            ticker = str(t).upper().strip()
            base_dados[ticker] = {
                'preco': limpar_valor(sheet.cell(row=row, column=COL_PRECO_DADOS).value),
                'dy':    limpar_valor(sheet.cell(row=row, column=COL_DY_DADOS).value),
                'pvp':   limpar_valor(sheet.cell(row=row, column=COL_PVP_DADOS).value),
            }
    wb.close()
    return base_dados, ultima_att


def carregar_carteira(base_b3):
    carteira   = {}
    patrimonio = 0.0
    sem_preco  = []
    wb    = openpyxl.load_workbook(ARQUIVO_EXCEL, data_only=True)
    sheet = wb[ABA_CARTEIRA]
    for row in range(LINHA_INICIO, sheet.max_row + 1):
        ativo = sheet.cell(row=row, column=COL_ATIVO_CART).value
        if not ativo or not isinstance(ativo, str):
            continue
        ticker = ativo.upper().strip()
        if ticker in ('ATIVO', 'TOTAL', 'TOTAL GERAL'):
            continue
        valor_atual = limpar_valor(sheet.cell(row=row, column=COL_VALOR_CART).value)
        qtde        = limpar_valor(sheet.cell(row=row, column=COL_QTDE_CART).value)
        pm          = limpar_valor(sheet.cell(row=row, column=COL_PM_CART).value)
        info        = base_b3.get(ticker, {})
        preco       = info.get('preco', 0.0)
        dy          = info.get('dy',    0.0)
        pvp         = info.get('pvp',   0.0)
        patrimonio += valor_atual
        if preco <= 0:
            sem_preco.append(ticker)
        carteira[ticker] = {
            'qtde':        qtde,
            'pm':          pm,
            'preco':       preco,
            'valor_atual': valor_atual,
            'dy':          dy,
            'pvp':         pvp,
        }
    wb.close()
    return carteira, patrimonio, sem_preco


# ══════════════════════════════════════════════════════════════════════
# MOTOR DE DECISÃO
# ══════════════════════════════════════════════════════════════════════

def calcular_pesos_alvo(ativos_com_preco: dict) -> dict:
    """
    Peso proporcional ao DY, com piso de 50 % do peso uniforme.
    Ativos sem DY recebem o peso mínimo.
    """
    n       = len(ativos_com_preco)
    soma_dy = sum(d['dy'] for d in ativos_com_preco.values())
    pesos   = {}
    if soma_dy > 0:
        peso_min = (1 / n) * 0.5
        for ticker, dados in ativos_com_preco.items():
            pesos[ticker] = dados['dy'] / soma_dy if dados['dy'] > 0 else peso_min
        total = sum(pesos.values())
        pesos = {t: p / total for t, p in pesos.items()}
    else:
        pesos = {t: 1 / n for t in ativos_com_preco}
    return pesos


def calcular_vendas(carteira: dict, patrimonio: float, pesos: dict) -> list:
    """
    Avalia cada ativo e sugere venda se um ou mais critérios forem atendidos.
    Retorna lista ordenada por urgência (score).
    """
    media_dy = (
        sum(d['dy'] for d in carteira.values() if d['preco'] > 0)
        / max(1, sum(1 for d in carteira.values() if d['preco'] > 0))
    )

    sugestoes = []
    for ticker, dados in carteira.items():
        preco       = dados['preco']
        qtde        = dados['qtde']
        pm          = dados['pm']
        dy          = dados['dy']
        pvp         = dados['pvp']
        valor_atual = dados['valor_atual']

        if preco <= 0 or qtde <= 0:
            continue

        motivos      = []
        urgencia     = 0.0
        pct_carteira = valor_atual / patrimonio if patrimonio > 0 else 0
        alvo         = pesos.get(ticker, 0)

        # ── 1. Sobrealinhamento ─────────────────────────────────────
        if alvo > 0 and pct_carteira > alvo * (1 + LIMIAR_SOBREPESO):
            excesso_pct = (pct_carteira / alvo - 1) * 100
            motivos.append(f'Sobrealinhado +{excesso_pct:.1f}% do alvo')
            urgencia += excesso_pct * 0.5

        # ── 2. DY muito baixo vs média da carteira ─────────────────
        if dy < DY_MINIMO and dy > 0 and media_dy > DY_MINIMO:
            deficit_dy = media_dy - dy
            motivos.append(f'DY baixo ({dy:.1f}% vs média {media_dy:.1f}%)')
            urgencia += deficit_dy * 2

        # ── 3. P/VP elevado (apenas FIIs) ─────────────────────────
        if e_fii(ticker) and pvp > PVP_MAXIMO_FII and pvp > 0:
            excesso_pvp = (pvp - 1) * 100
            motivos.append(f'P/VP elevado ({pvp:.2f}x — paga {excesso_pvp:.0f}% sobre o PL)')
            urgencia += (pvp - PVP_MAXIMO_FII) * 30

        # ── 4. Stop loss ────────────────────────────────────────────
        if pm > 0 and preco < pm * (1 - STOP_PERDA):
            perda_pct = (1 - preco / pm) * 100
            motivos.append(f'Stop loss: queda de {perda_pct:.1f}% sobre PM')
            urgencia += perda_pct * 1.5

        # ── 5. Take profit (realização parcial) ────────────────────
        if pm > 0 and preco > pm * (1 + TAKE_PROFIT):
            ganho_pct = (preco / pm - 1) * 100
            motivos.append(f'Take profit: ganho de {ganho_pct:.1f}% sobre PM')
            urgencia += ganho_pct * 0.8

        if not motivos:
            continue

        # Quantidade sugerida de venda (fração da posição, mínimo 1 cota)
        cotas_sugeridas = max(MIN_COTAS_VENDA,
                              math.floor(qtde * FRACAO_VENDA))
        # Não sugere mais do que a posição inteira
        cotas_sugeridas = min(cotas_sugeridas, int(qtde))
        valor_venda     = cotas_sugeridas * preco

        sugestoes.append({
            'ativo':           ticker,
            'preco':           preco,
            'qtde_atual':      qtde,
            'pm':              pm,
            'dy':              dy,
            'pvp':             pvp,
            'pct_carteira':    pct_carteira * 100,
            'alvo_pct':        alvo * 100,
            'motivos':         motivos,
            'urgencia':        urgencia,
            'cotas_vender':    cotas_sugeridas,
            'valor_venda':     valor_venda,
        })

    sugestoes.sort(key=lambda x: x['urgencia'], reverse=True)
    return sugestoes


def calcular_aporte(carteira: dict, patrimonio: float,
                    valor_aporte: float, tickers_venda: set) -> tuple:
    """
    Calcula a boleta de compra ignorando ativos com sugestão de venda.
    Score = défice * (1 + DY/100) * fator_pvp
    fator_pvp: penaliza ativos com P/VP > 1 (FIIs caros)
    """
    ativos_com_preco = {
        t: d for t, d in carteira.items()
        if d['preco'] > 0 and t not in tickers_venda
    }
    if not ativos_com_preco:
        return [], valor_aporte

    patrimonio_proj = patrimonio + valor_aporte
    pesos           = calcular_pesos_alvo(ativos_com_preco)

    analise = []
    for ticker, dados in ativos_com_preco.items():
        alvo   = patrimonio_proj * pesos[ticker]
        defice = alvo - dados['valor_atual']
        if defice <= 0:
            continue

        pvp       = dados['pvp']
        fator_pvp = 1.0
        if e_fii(ticker) and pvp > 1.0:
            fator_pvp = max(0.5, 1 / pvp)  # penaliza FIIs caros

        score = defice * (1 + dados['dy'] / 100) * fator_pvp

        analise.append({
            'ativo':     ticker,
            'preco':     dados['preco'],
            'defice':    defice,
            'dy':        dados['dy'],
            'pvp':       pvp,
            'score':     score,
            'peso_alvo': pesos[ticker] * 100,
            'alvo_rs':   alvo,
        })

    analise.sort(key=lambda x: x['score'], reverse=True)

    compras = []
    caixa   = valor_aporte
    for item in analise:
        preco         = item['preco']
        cotas_comprar = min(
            math.floor(item['defice'] / preco),
            math.floor(caixa / preco)
        )
        if cotas_comprar > 0:
            valor_ordem = cotas_comprar * preco
            compras.append({
                'ativo':       item['ativo'],
                'qtde':        cotas_comprar,
                'preco':       preco,
                'total_ordem': valor_ordem,
                'dy':          item['dy'],
                'pvp':         item['pvp'],
                'peso_alvo':   item['peso_alvo'],
                'alvo_rs':     item['alvo_rs'],
            })
            caixa -= valor_ordem

    return compras, caixa


# ══════════════════════════════════════════════════════════════════════
# INTERFACE GRÁFICA
# ══════════════════════════════════════════════════════════════════════

class SmartAporteApp:

    def __init__(self, root):
        self.root = root
        self.root.title('Smart Aporte Pro — Compra & Venda')
        self.root.resizable(True, True)
        self.root.minsize(820, 620)
        self._centralizar(900, 700)
        self._construir_ui()

    # ------------------------------------------------------------------
    # Layout principal
    # ------------------------------------------------------------------
    def _construir_ui(self):
        self.root.configure(bg=COR_CINZA_BG)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(2, weight=1)

        # ── Cabeçalho ─────────────────────────────────────────────────
        header = tk.Frame(self.root, bg=COR_VERDE, padx=20, pady=14)
        header.grid(row=0, column=0, sticky='ew')
        header.columnconfigure(0, weight=1)

        tk.Label(header, text='Smart Aporte Pro — Compra & Venda',
                 bg=COR_VERDE, fg='white',
                 font=('Segoe UI', 14, 'bold'), anchor='w'
                 ).grid(row=0, column=0, sticky='w')
        tk.Label(header,
                 text='Rebalanceamento inteligente com critérios de mercado B3',
                 bg=COR_VERDE, fg='#a7f3d0',
                 font=('Segoe UI', 9), anchor='w'
                 ).grid(row=1, column=0, sticky='w')

        # ── Painel de entrada ─────────────────────────────────────────
        entrada = tk.Frame(self.root, bg=COR_CINZA_BG, padx=20, pady=12)
        entrada.grid(row=1, column=0, sticky='ew')

        tk.Label(entrada, text='Valor do aporte (R$):',
                 bg=COR_CINZA_BG, font=('Segoe UI', 10)
                 ).grid(row=0, column=0, sticky='w')

        self.var_valor = tk.StringVar()
        entry = tk.Entry(entrada, textvariable=self.var_valor,
                         font=('Segoe UI', 13, 'bold'), width=16, justify='right')
        entry.grid(row=0, column=1, padx=(10, 0), sticky='w')
        entry.bind('<Return>', lambda e: self._executar())
        entry.focus()

        self.btn_calcular = tk.Button(
            entrada, text='⚡  Calcular boleta',
            command=self._executar,
            bg=COR_VERDE, fg='white',
            font=('Segoe UI', 10, 'bold'),
            padx=18, pady=6, cursor='hand2',
            relief='flat', activebackground=COR_VERDE_ESC
        )
        self.btn_calcular.grid(row=0, column=2, padx=(20, 0))

        self.var_status_cot = tk.StringVar(value='')
        self.lbl_cot = tk.Label(entrada, textvariable=self.var_status_cot,
                                bg=COR_CINZA_BG, font=('Segoe UI', 8), fg='gray')
        self.lbl_cot.grid(row=1, column=0, columnspan=4, sticky='w', pady=(4, 0))

        # ── Área com abas Compra / Venda ───────────────────────────────
        nb_frame = tk.Frame(self.root, bg=COR_CINZA_BG, padx=14, pady=0)
        nb_frame.grid(row=2, column=0, sticky='nsew', pady=(0, 4))
        nb_frame.columnconfigure(0, weight=1)
        nb_frame.rowconfigure(1, weight=1)

        # Cards de resumo acima das abas
        self._construir_cards(nb_frame)

        # Notebook com abas
        style = ttk.Style()
        style.configure('TNotebook',       background=COR_CINZA_BG, borderwidth=0)
        style.configure('TNotebook.Tab',   font=('Segoe UI', 10, 'bold'), padding=[14, 6])
        style.map('TNotebook.Tab',
                  background=[('selected', COR_VERDE)],
                  foreground=[('selected', 'white')])

        self.nb = ttk.Notebook(nb_frame)
        self.nb.grid(row=1, column=0, sticky='nsew')

        # Aba COMPRA
        aba_compra = tk.Frame(self.nb, bg='white')
        self.nb.add(aba_compra, text='  📈  Comprar  ')
        self._construir_tabela_compra(aba_compra)

        # Aba VENDA
        aba_venda = tk.Frame(self.nb, bg='white')
        self.nb.add(aba_venda, text='  📉  Vender  ')
        self._construir_tabela_venda(aba_venda)

        # Rodapé
        self.var_msg = tk.StringVar(
            value='Informe o valor do aporte e pressione Enter ou clique em "Calcular boleta".')
        tk.Label(self.root, textvariable=self.var_msg,
                 bg=COR_CINZA_BG, font=('Segoe UI', 9), fg='gray', anchor='w'
                 ).grid(row=3, column=0, sticky='ew', padx=20, pady=(2, 10))

        self._atualizar_status_cotacoes()

    def _construir_cards(self, parent):
        cards = tk.Frame(parent, bg=COR_CINZA_BG)
        cards.grid(row=0, column=0, sticky='ew', pady=(0, 8))
        for i in range(5):
            cards.columnconfigure(i, weight=1)

        self.lbl_patrimonio  = self._card(cards, 'Patrimônio',    '—', 0, COR_AZUL)
        self.lbl_ordens_c    = self._card(cards, 'Ordens compra', '—', 1, COR_VERDE)
        self.lbl_alocado     = self._card(cards, 'Total alocado', '—', 2, COR_VERDE)
        self.lbl_caixa       = self._card(cards, 'Caixa restante','—', 3, COR_AMARELO)
        self.lbl_ordens_v    = self._card(cards, 'Sugestões venda','—', 4, COR_VERM)

    def _card(self, parent, titulo, valor, col, cor_topo):
        frame = tk.Frame(parent, bg='white',
                         highlightthickness=1,
                         highlightbackground='#e0e0da')
        frame.grid(row=0, column=col, sticky='nsew', padx=(0, 6))
        topo = tk.Frame(frame, bg=cor_topo, height=4)
        topo.pack(fill='x')
        inner = tk.Frame(frame, bg='white', padx=10, pady=8)
        inner.pack(fill='x')
        tk.Label(inner, text=titulo, bg='white',
                 font=('Segoe UI', 8), fg='gray').pack(anchor='w')
        var = tk.StringVar(value=valor)
        tk.Label(inner, textvariable=var, bg='white',
                 font=('Segoe UI', 11, 'bold')).pack(anchor='w')
        return var

    def _construir_tabela_compra(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(0, weight=1)

        cols = ('ativo', 'qtde', 'preco', 'dy', 'pvp', 'peso_alvo', 'alvo_rs', 'total')
        self.tree_c = ttk.Treeview(parent, columns=cols, show='headings', height=12)

        defs = [
            ('ativo',    'Ativo',   70,  'center'),
            ('qtde',     'Qtde',    55,  'center'),
            ('preco',    'Preço',   88,  'e'),
            ('dy',       'DY %',    68,  'center'),
            ('pvp',      'P/VP',    68,  'center'),
            ('peso_alvo','Alvo %',  68,  'center'),
            ('alvo_rs',  'Alvo R$', 100, 'e'),
            ('total',    'Total',   100, 'e'),
        ]
        for col, titulo, larg, ancora in defs:
            self.tree_c.heading(col, text=titulo)
            self.tree_c.column(col, width=larg, anchor=ancora, stretch=False)

        self.tree_c.tag_configure('par',   background='#f9f9f6')
        self.tree_c.tag_configure('impar', background='#ffffff')

        scroll_c = ttk.Scrollbar(parent, orient='vertical', command=self.tree_c.yview)
        self.tree_c.configure(yscrollcommand=scroll_c.set)
        self.tree_c.grid(row=0, column=0, sticky='nsew')
        scroll_c.grid(row=0, column=1, sticky='ns')

    def _construir_tabela_venda(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(0, weight=1)

        cols = ('ativo', 'cotas', 'preco', 'pm', 'dy', 'pvp',
                'pct_cart', 'alvo_pct', 'valor_venda', 'motivo')
        self.tree_v = ttk.Treeview(parent, columns=cols, show='headings', height=12)

        defs = [
            ('ativo',      'Ativo',      70,  'center'),
            ('cotas',      'Vender',     60,  'center'),
            ('preco',      'Preço',      84,  'e'),
            ('pm',         'PM',         84,  'e'),
            ('dy',         'DY %',       64,  'center'),
            ('pvp',        'P/VP',       64,  'center'),
            ('pct_cart',   'Peso %',     64,  'center'),
            ('alvo_pct',   'Alvo %',     64,  'center'),
            ('valor_venda','Vl. Venda',  100, 'e'),
            ('motivo',     'Motivo principal', 200, 'w'),
        ]
        for col, titulo, larg, ancora in defs:
            self.tree_v.heading(col, text=titulo)
            self.tree_v.column(col, width=larg, anchor=ancora, stretch=False)

        self.tree_v.tag_configure('venda_par',   background='#fff5f5')
        self.tree_v.tag_configure('venda_impar', background='#fff9f9')
        self.tree_v.tag_configure('stop',        background='#fee2e2', foreground='#7f1d1d')
        self.tree_v.tag_configure('take',        background='#dcfce7', foreground='#14532d')

        scroll_v = ttk.Scrollbar(parent, orient='vertical', command=self.tree_v.yview)
        self.tree_v.configure(yscrollcommand=scroll_v.set)
        self.tree_v.grid(row=0, column=0, sticky='nsew')
        scroll_v.grid(row=0, column=1, sticky='ns')

        # Legenda abaixo da tabela de venda
        legenda = tk.Frame(parent, bg='white', padx=10, pady=6)
        legenda.grid(row=1, column=0, columnspan=2, sticky='ew')
        for cor, texto in [('#fee2e2', 'Stop loss'), ('#dcfce7', 'Take profit'),
                           ('#fff5f5', 'Rebalanceamento / P/VP / DY')]:
            dot = tk.Frame(legenda, bg=cor, width=14, height=14,
                           highlightthickness=1, highlightbackground='#ccc')
            dot.pack(side='left', padx=(0, 4))
            tk.Label(legenda, text=texto, bg='white',
                     font=('Segoe UI', 8), fg='gray').pack(side='left', padx=(0, 14))

    # ------------------------------------------------------------------
    # Execução
    # ------------------------------------------------------------------
    def _atualizar_status_cotacoes(self):
        try:
            mtime = datetime.fromtimestamp(os.path.getmtime(ARQUIVO_EXCEL))
            delta = datetime.now() - mtime
            horas = int(delta.total_seconds() // 3600)
            if horas > 24:
                self.var_status_cot.set(
                    f'⚠  Cotações atualizadas há {horas}h — rode "Atualizar Cotações" primeiro.')
                self.lbl_cot.config(fg=COR_AMARELO)
            else:
                self.var_status_cot.set(
                    f'✓  Cotações atualizadas em {mtime.strftime("%d/%m/%Y %H:%M")}')
                self.lbl_cot.config(fg='gray')
        except Exception:
            pass

    def _executar(self):
        raw = self.var_valor.get().strip().replace('.', '').replace(',', '.')
        try:
            valor_aporte = float(raw)
            if valor_aporte <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror('Valor inválido',
                                 'Digite um valor numérico positivo.\nExemplo: 1500,00')
            return

        self.btn_calcular.config(state='disabled', text='Calculando...')
        self.var_msg.set('Lendo planilha e calculando boleta...')
        self.root.update()
        threading.Thread(target=self._processar,
                         args=(valor_aporte,), daemon=True).start()

    def _processar(self, valor_aporte):
        try:
            base_b3, _                    = carregar_dados_b3()
            carteira, patrimonio, sem_preco = carregar_carteira(base_b3)

            # Pesos alvo para toda a carteira (usado em venda)
            ativos_todos = {t: d for t, d in carteira.items() if d['preco'] > 0}
            pesos_todos  = calcular_pesos_alvo(ativos_todos) if ativos_todos else {}

            vendas        = calcular_vendas(carteira, patrimonio, pesos_todos)
            tickers_venda = {v['ativo'] for v in vendas}
            compras, caixa = calcular_aporte(carteira, patrimonio,
                                             valor_aporte, tickers_venda)

            self.root.after(
                0, self._exibir,
                patrimonio, sem_preco, compras, caixa, vendas
            )
        except Exception as e:
            import traceback
            msg = (
                f"{e}\n\n--- Diagnóstico ---\n"
                f"Excel: {ARQUIVO_EXCEL}\n"
                f"Excel existe: {os.path.isfile(ARQUIVO_EXCEL)}\n\n"
                f"{traceback.format_exc()[-400:]}"
            )
            self.root.after(0, self._exibir_erro, msg)

    def _exibir(self, patrimonio, sem_preco, compras, caixa, vendas):
        total_alocado = sum(o['total_ordem'] for o in compras)

        self.lbl_patrimonio.set(f'R$ {patrimonio:,.2f}')
        self.lbl_ordens_c.set(str(len(compras)) if compras else '—')
        self.lbl_alocado.set(f'R$ {total_alocado:,.2f}' if compras else '—')
        self.lbl_caixa.set(f'R$ {caixa:,.2f}')
        self.lbl_ordens_v.set(str(len(vendas)) if vendas else '—')

        # ── Tabela de compra ──────────────────────────────────────────
        for row in self.tree_c.get_children():
            self.tree_c.delete(row)

        for i, o in enumerate(compras):
            pvp_txt = f"{o['pvp']:.2f}x" if o['pvp'] > 0 else '—'
            self.tree_c.insert('', 'end',
                tags=('par' if i % 2 == 0 else 'impar',),
                values=(
                    o['ativo'],
                    f"{o['qtde']}x",
                    f"R$ {o['preco']:.2f}",
                    f"{o['dy']:.2f}%",
                    pvp_txt,
                    f"{o['peso_alvo']:.1f}%",
                    f"R$ {o['alvo_rs']:.2f}",
                    f"R$ {o['total_ordem']:.2f}",
                ))

        # ── Tabela de venda ───────────────────────────────────────────
        for row in self.tree_v.get_children():
            self.tree_v.delete(row)

        for i, v in enumerate(vendas):
            motivo_principal = v['motivos'][0]
            pm_txt  = f"R$ {v['pm']:.2f}"  if v['pm']  > 0 else '—'
            pvp_txt = f"{v['pvp']:.2f}x"   if v['pvp'] > 0 else '—'

            # Escolhe tag por tipo de motivo
            m = motivo_principal.lower()
            if 'stop' in m:
                tag = 'stop'
            elif 'take' in m or 'profit' in m or 'ganho' in m:
                tag = 'take'
            elif i % 2 == 0:
                tag = 'venda_par'
            else:
                tag = 'venda_impar'

            self.tree_v.insert('', 'end',
                tags=(tag,),
                values=(
                    v['ativo'],
                    f"{v['cotas_vender']}x",
                    f"R$ {v['preco']:.2f}",
                    pm_txt,
                    f"{v['dy']:.2f}%",
                    pvp_txt,
                    f"{v['pct_carteira']:.1f}%",
                    f"{v['alvo_pct']:.1f}%",
                    f"R$ {v['valor_venda']:.2f}",
                    motivo_principal,
                ))

        # Muda aba automaticamente se houver vendas urgentes
        if vendas:
            self.nb.select(1)

        # Aviso sobre ativos ignorados na compra
        ignorados = len({v['ativo'] for v in vendas})
        avisos = []
        if ignorados:
            avisos.append(f'{ignorados} ativo(s) com sugestão de venda excluído(s) da compra')
        if sem_preco:
            avisos.append(f'{len(sem_preco)} ativo(s) sem cotação ignorado(s)')

        n_compras = len(compras)
        n_vendas  = len(vendas)
        partes    = [f'{n_compras} ordem(s) de compra']
        if n_vendas:
            partes.append(f'{n_vendas} sugestão(ões) de venda')
        msg = '  |  '.join(partes)
        if avisos:
            msg += '   ⚠  ' + ' · '.join(avisos)

        self.var_msg.set(msg)
        self.btn_calcular.config(state='normal', text='⚡  Calcular boleta')

    def _exibir_erro(self, mensagem):
        self.var_msg.set(f'Erro: {mensagem[:120]}...')
        self.btn_calcular.config(state='normal', text='⚡  Calcular boleta')
        messagebox.showerror('Erro ao processar', mensagem)

    def _centralizar(self, largura, altura):
        self.root.update_idletasks()
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        self.root.geometry(
            f'{largura}x{altura}+{(sw - largura) // 2}+{(sh - altura) // 2}')


# ══════════════════════════════════════════════════════════════════════

def main():
    root = tk.Tk()
    SmartAporteApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()
