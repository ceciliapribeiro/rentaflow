import openpyxl
import requests
import shutil
from datetime import datetime, timedelta
from collections import defaultdict
import time
import sys
import os
import threading
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config_loader import carregar_config, fazer_backup

_cfg              = carregar_config()
ARQUIVO_EXCEL     = _cfg['arquivo_excel']
JANELA_BUSCA_DIAS = _cfg['janela_busca_dias']

ABA_CARTEIRA   = 'CARTEIRA'
ABA_DIVIDENDOS = 'DIVIDENDOS'
ABA_DADOS      = 'Dados B3'
ABA_OPERACOES  = 'OPERAÇÕES'

COL_OP_DATA   = 1
COL_OP_TICKER = 2
COL_OP_QTDE   = 3
COL_OP_TIPO   = 8

COR_VERDE     = '#1a6b45'
COR_VERDE_ESC = '#145535'
COR_AMARELO   = '#b45309'
COR_ERRO      = '#991b1b'


# ══════════════════════════════════════════════════════════════════════
# LÓGICA DE NEGÓCIO (sem prints — usa callback para a UI)
# ══════════════════════════════════════════════════════════════════════

def _adicionar_dias_uteis(data, dias):
    """
    Avança 'dias' dias úteis a partir de 'data', pulando sábados e domingos.
    Usado para calcular a data de liquidação (D+2) das operações.
    """
    atual = data
    adicionados = 0
    while adicionados < dias:
        atual = atual.replace(hour=0, minute=0, second=0, microsecond=0)
        atual = atual + __import__('datetime').timedelta(days=1)
        if atual.weekday() < 5:   # 0=seg ... 4=sex
            adicionados += 1
    return atual


def construir_historico_posicoes():
    """
    Reconstrói o histórico de custódia por ticker.
    Usa a data de LIQUIDAÇÃO (D+2 úteis) em vez da data do pregão,
    pois a B3 só considera as cotas elegíveis a proventos após a liquidação.
    Isso evita contabilizar cotas compradas no mesmo dia (ou após) a data EX.
    """
    wb    = openpyxl.load_workbook(ARQUIVO_EXCEL, data_only=True)
    sheet = wb[ABA_OPERACOES]
    ops_por_ticker = defaultdict(list)

    for row in range(2, sheet.max_row + 1):
        data_raw = sheet.cell(row=row, column=COL_OP_DATA).value
        ticker   = sheet.cell(row=row, column=COL_OP_TICKER).value
        qtde_raw = sheet.cell(row=row, column=COL_OP_QTDE).value
        tipo_raw = sheet.cell(row=row, column=COL_OP_TIPO).value

        if not ticker or not data_raw or not tipo_raw:
            continue
        if isinstance(data_raw, datetime):
            data_pregao = data_raw.replace(hour=0, minute=0, second=0, microsecond=0)
        else:
            try:
                data_pregao = datetime.strptime(str(data_raw).split(' ')[0], '%Y-%m-%d')
            except:
                continue
        try:
            qtde = float(qtde_raw or 0)
        except:
            continue

        tipo   = str(tipo_raw).strip().upper()
        ticker = str(ticker).strip().upper()

        # Liquidação D+2 úteis: as cotas só estão na custódia após este prazo
        data_liquidacao = _adicionar_dias_uteis(data_pregao, 2)

        if 'COMPRA' in tipo:
            ops_por_ticker[ticker].append((data_liquidacao, +qtde))
        elif 'VENDA' in tipo:
            ops_por_ticker[ticker].append((data_liquidacao, -qtde))

    historico = {}
    for ticker, ops in ops_por_ticker.items():
        ops.sort(key=lambda x: x[0])
        acumulado = 0.0
        snapshots = []
        for data_op, delta in ops:
            acumulado += delta
            snapshots.append((data_op, max(acumulado, 0)))
        historico[ticker] = snapshots

    wb.close()
    return historico


def qtde_em_custodia_na_data(snapshots, data_referencia):
    qtde = 0.0
    for data_op, qtde_acumulada in snapshots:
        if data_op <= data_referencia:
            qtde = qtde_acumulada
        else:
            break
    return qtde


def extrair_dados_b3(wb):
    sheet = wb[ABA_DADOS]
    dados = {}
    for row in range(2, sheet.max_row + 1):
        ativo = sheet.cell(row=row, column=1).value
        tipo  = sheet.cell(row=row, column=2).value
        razao = sheet.cell(row=row, column=3).value
        if ativo:
            dados[str(ativo).upper().strip()] = {'tipo': tipo, 'razao': razao}
    return dados


def ler_dividendos_existentes(sheet):
    """
    Retorna dict {chave: {'valor': float, 'linha': int}} para cada
    provento já registrado.
    Chave: "TICKER_DD/MM/AAAA_TIPO" (TIPO = RENDIMENTO ou JUROS)
    Inclui o tipo para evitar colisão entre RENDIMENTO e JUROS
    pagos na mesma data pelo mesmo ativo.
    """
    existentes = {}
    for row in range(2, sheet.max_row + 1):
        ativo   = sheet.cell(row=row, column=7).value
        data_op = sheet.cell(row=row, column=2).value
        valor   = sheet.cell(row=row, column=5).value
        tipo    = sheet.cell(row=row, column=6).value
        if not ativo or not data_op:
            continue
        try:
            if isinstance(data_op, datetime):
                data_str = data_op.strftime('%d/%m/%Y')
            else:
                data_str = str(data_op).split(' ')[0]
                if '-' in data_str and len(data_str.split('-')[0]) == 4:
                    partes   = data_str.split('-')
                    data_str = f"{partes[2]}/{partes[1]}/{partes[0]}"
            tipo_str = str(tipo).upper().strip() if tipo else 'RENDIMENTO'
            chave = f"{str(ativo).upper().strip()}_{data_str}_{tipo_str}"
            existentes[chave] = {
                'valor': float(valor) if valor is not None else None,
                'linha': row,
            }
        except:
            pass
    return existentes


def _tipo_para_desc(tipo_raw):
    """Normaliza o tipo de provento para RENDIMENTO ou JUROS."""
    t = str(tipo_raw).upper()
    if any(x in t for x in ('JRS', 'JUROS', 'JCP', 'CAP PROPRIO')):
        return 'JUROS'
    return 'RENDIMENTO'


def obter_proventos_fundamentus(ticker, tipo_ativo):
    """
    Busca proventos no Fundamentus, que usa dados diretos da B3.
    FIIs  → fii_proventos.php  : 4 campos/linha (data_ex, tipo, data_pgto, valor)
    Ações → proventos.php      : 5 campos/linha (data_ex, valor, tipo, data_pgto, _)
    Retorna lista de dicts com data_ex, data_pagamento, valor, tipo.
    """
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Referer':    'https://www.fundamentus.com.br/',
    }
    eh_fii = tipo_ativo and str(tipo_ativo).strip().upper() in ('FII', 'FIAGRO')

    if eh_fii:
        url = f'https://www.fundamentus.com.br/fii_proventos.php?papel={ticker.upper()}'
    else:
        url = f'https://www.fundamentus.com.br/proventos.php?papel={ticker.upper()}&tipo=2'

    for tentativa in range(3):
        try:
            req = requests.get(url, headers=headers, timeout=10)
            if req.status_code == 200:
                break
            if tentativa < 2:
                time.sleep(3 * (tentativa + 1))
        except Exception:
            if tentativa < 2:
                time.sleep(3 * (tentativa + 1))
    else:
        return []

    import re
    campos = [c.strip() for c in re.findall(r'<td[^>]*>([^<]+)</td>', req.text)]
    if not campos:
        return []

    proventos = []
    passo = 4 if eh_fii else 5

    for i in range(0, len(campos) - (passo - 1), passo):
        try:
            if eh_fii:
                # FII: data_ex | tipo | data_pagamento | valor
                data_ex_str  = campos[i].strip()
                tipo_raw     = campos[i + 1].strip()
                data_pag_str = campos[i + 2].strip().split()[0]
                valor_str    = campos[i + 3].strip()
            else:
                # Ação: data_ex | valor | tipo | data_pagamento | _
                data_ex_str  = campos[i].strip()
                valor_str    = campos[i + 1].strip()
                tipo_raw     = campos[i + 2].strip()
                data_pag_str = campos[i + 3].strip().split()[0]

            # Ignora linhas sem data de pagamento definida
            if not data_pag_str or data_pag_str == '-':
                continue

            data_ex  = datetime.strptime(data_ex_str,  '%d/%m/%Y')
            data_pag = datetime.strptime(data_pag_str, '%d/%m/%Y')
            valor    = float(valor_str.replace('.', '').replace(',', '.'))

            proventos.append({
                'data_ex':        data_ex,
                'data_pagamento': data_pag,
                'valor':          valor,
                'tipo':           _tipo_para_desc(tipo_raw),
            })
        except Exception:
            continue

    return proventos


def obter_proventos_status_invest(ticker, tipo_ativo):
    """Fallback: Status Invest — usado quando Fundamentus não retorna dados."""
    headers = {'User-Agent': 'Mozilla/5.0'}
    eh_fii  = tipo_ativo and str(tipo_ativo).strip().upper() == 'FII'
    if eh_fii:
        url = (f"https://statusinvest.com.br/fii/companytickerprovents"
               f"?ticker={ticker.lower()}&chartProventsType=2")
    else:
        url = (f"https://statusinvest.com.br/acao/companytickerprovents"
               f"?ticker={ticker.lower()}&chartProventsType=1")

    def _buscar(u):
        for tentativa in range(3):
            try:
                req = requests.get(u, headers=headers, timeout=10)
                if req.status_code in (429, 503):
                    time.sleep(15 * (tentativa + 1))
                    continue
                if req.status_code == 200:
                    return req.json().get('assetEarningsModels', [])
            except Exception:
                if tentativa < 2:
                    time.sleep(3 * (tentativa + 1))
        return []

    modelos = _buscar(url)
    if not modelos and not eh_fii:
        url_alt = url.replace('chartProventsType=1', 'chartProventsType=2')
        modelos = _buscar(url_alt)

    proventos = []
    for item in modelos:
        if not item.get('pd') or item.get('pd') == '-':
            continue
        if not item.get('ed') or item.get('ed') == '-':
            continue
        try:
            proventos.append({
                'data_ex':        datetime.strptime(item['ed'], '%d/%m/%Y'),
                'data_pagamento': datetime.strptime(item['pd'], '%d/%m/%Y'),
                'valor':          float(item['v']),
                'tipo':           _tipo_para_desc(item.get('et', 'Rendimento')),
            })
        except Exception:
            continue
    return proventos


def obter_proventos(ticker, tipo_ativo, log_cb=None):
    """
    Busca proventos usando Fundamentus como fonte primária (dados B3 diretos)
    e Status Invest como fallback.
    """
    proventos = obter_proventos_fundamentus(ticker, tipo_ativo)

    if not proventos:
        if log_cb:
            log_cb(f'  [AVISO] {ticker}: Fundamentus sem dados — tentando Status Invest...', 'aviso')
        proventos = obter_proventos_status_invest(ticker, tipo_ativo)

    return proventos



def encontrar_primeira_linha_vazia(sheet):
    for row in range(2, sheet.max_row + 100):
        if sheet.cell(row=row, column=7).value is None:
            return row
    return sheet.max_row + 1


def rastrear_e_inserir_dividendos(log_cb, progresso_cb, parar_evento):
    """
    Versão sem prints — usa callbacks para comunicar com a UI.
    log_cb(msg, tag)  : envia linha ao log visual ('info', 'novo', 'aviso', 'erro')
    progresso_cb(i, n): atualiza a barra de progresso
    parar_evento      : threading.Event — checado a cada ticker
    Retorna dict com estatísticas finais.
    """
    stats = {'inseridos': 0, 'ignorados': 0, 'erros': 0, 'tickers': 0}

    # Passo 1 — histórico
    log_cb('Reconstruindo histórico de posições...', 'info')
    try:
        historico = construir_historico_posicoes()
    except Exception as e:
        log_cb(f'Erro ao ler OPERAÇÕES: {e}', 'erro')
        raise

    if not historico:
        log_cb('Nenhuma operação encontrada na aba OPERAÇÕES.', 'aviso')
        return stats

    stats['tickers'] = len(historico)
    log_cb(f'{len(historico)} ativos identificados no histórico.', 'info')

    # Passo 2 — carrega planilha e dividendos existentes
    log_cb('Mapeando proventos já registrados...', 'info')
    try:
        fazer_backup(ARQUIVO_EXCEL, log_cb)

        wb_escrita = openpyxl.load_workbook(ARQUIVO_EXCEL)
        meta_dados = extrair_dados_b3(wb_escrita)
        sheet_div  = wb_escrita[ABA_DIVIDENDOS]
        dividendos_existentes = ler_dividendos_existentes(sheet_div)
    except Exception as e:
        log_cb(f'Erro ao abrir planilha: {e}', 'erro')
        raise

    hoje_limite  = datetime.now().replace(hour=23, minute=59, second=59)
    inicio_busca = hoje_limite - timedelta(days=JANELA_BUSCA_DIAS)
    log_cb(
        f'Janela: {inicio_busca.strftime("%d/%m/%Y")} → {hoje_limite.strftime("%d/%m/%Y")}',
        'info'
    )

    linha_insercao = encontrar_primeira_linha_vazia(sheet_div)
    tickers_lista  = sorted(historico.items())
    total          = len(tickers_lista)

    # Passo 3 — loop principal
    for i, (ticker, snapshots) in enumerate(tickers_lista, 1):
        if parar_evento.is_set():
            log_cb('Interrompido pelo usuário.', 'aviso')
            break

        progresso_cb(i, total)
        tipo_ativo = meta_dados.get(ticker, {}).get('tipo', 'Ação/FII')

        try:
            proventos = obter_proventos(ticker, tipo_ativo, log_cb)
        except Exception as e:
            log_cb(f'[ERRO] {ticker}: {e}', 'erro')
            stats['erros'] += 1
            continue

        for prov in proventos:
            data_pagamento = prov['data_pagamento']
            if not (inicio_busca <= data_pagamento <= hoje_limite):
                continue

            qtde = qtde_em_custodia_na_data(snapshots, prov['data_ex'])
            if qtde <= 0:
                stats['ignorados'] += 1
                continue

            data_pag_str  = data_pagamento.strftime('%d/%m/%Y')
            tipo_original = prov['tipo'].upper()
            provento_desc = ('JUROS'
                             if 'JCP' in tipo_original or 'JUROS' in tipo_original
                             else 'RENDIMENTO')
            valor_total   = round(qtde * prov['valor'], 2)
            chave         = f"{ticker}_{data_pag_str}_{provento_desc}"

            if chave in dividendos_existentes:
                valor_registrado = dividendos_existentes[chave]['valor']
                if valor_registrado is not None and abs(valor_registrado - valor_total) > 0.01:
                    linha_corr = dividendos_existentes[chave]['linha']
                    sheet_div.cell(row=linha_corr, column=5).value = valor_total
                    log_cb(
                        f'[CORRIGIDO] {ticker:<8} {provento_desc:<12} '
                        f'Pgto: {data_pag_str}  '
                        f'R$ {valor_registrado:.2f} → R$ {valor_total:.2f} '
                        f'({qtde:.0f} cotas × R$ {prov["valor"]:.4f})',
                        'aviso'
                    )
                    stats['inseridos'] += 1
                continue

            sheet_div.cell(row=linha_insercao, column=1).value = data_pagamento.year
            sheet_div.cell(row=linha_insercao, column=2).value = data_pagamento   # datetime object
            sheet_div.cell(row=linha_insercao, column=5).value = valor_total
            sheet_div.cell(row=linha_insercao, column=6).value = provento_desc
            sheet_div.cell(row=linha_insercao, column=7).value = ticker

            log_cb(
                f'[NOVO]  {ticker:<8} {provento_desc:<12} '
                f'{qtde:.0f} cotas × R$ {prov["valor"]:.4f} = R$ {valor_total:.2f}'
                f'   Pgto: {data_pag_str}',
                'novo'
            )
            dividendos_existentes[chave] = {'valor': valor_total, 'linha': linha_insercao}
            linha_insercao   += 1
            stats['inseridos'] += 1

        time.sleep(0.5)

    # Passo 4 — salva
    try:
        if stats['inseridos'] > 0:
            log_cb(f'Salvando {stats["inseridos"]} novos registros...', 'info')
            wb_escrita.save(ARQUIVO_EXCEL)
            log_cb('Arquivo salvo com sucesso.', 'info')
        else:
            log_cb('Nenhum provento novo encontrado.', 'aviso')
    finally:
        wb_escrita.close()

    return stats


# ══════════════════════════════════════════════════════════════════════
# INTERFACE GRÁFICA
# ══════════════════════════════════════════════════════════════════════

class CacaDividendosApp:
    def __init__(self, root):
        self.root        = root
        self.parar_evt   = threading.Event()
        self.em_execucao = False

        self.root.title('Buscar Dividendos — Caça Proventos')
        self.root.resizable(False, False)
        self._centralizar(700, 580)
        self._construir_ui()

    # ------------------------------------------------------------------
    # Layout
    # ------------------------------------------------------------------
    def _construir_ui(self):
        # Cabeçalho
        header = tk.Frame(self.root, bg=COR_VERDE, height=60)
        header.pack(fill='x')
        header.pack_propagate(False)
        tk.Label(header, text='  Buscar Dividendos — Caça Proventos',
                 bg=COR_VERDE, fg='white',
                 font=('Segoe UI', 13, 'bold'), anchor='w'
                 ).pack(fill='both', expand=True, padx=4)

        # Cards de resumo
        resumo = tk.Frame(self.root, bg='#f5f5f0', padx=20, pady=10)
        resumo.pack(fill='x')
        self.card_tickers  = self._card(resumo, 'Ativos consultados', '—', 0)
        self.card_inseridos = self._card(resumo, 'Novos proventos',    '—', 1)
        self.card_ignorados = self._card(resumo, 'Sem cotas na data',  '—', 2)
        self.card_erros     = self._card(resumo, 'Erros de consulta',  '—', 3)
        resumo.columnconfigure((0, 1, 2, 3), weight=1)

        # Barra de progresso
        prog_frame = tk.Frame(self.root, padx=20, pady=6)
        prog_frame.pack(fill='x')

        self.var_prog_label = tk.StringVar(value='Aguardando...')
        tk.Label(prog_frame, textvariable=self.var_prog_label,
                 font=('Segoe UI', 8), fg='gray', anchor='w'
                 ).pack(fill='x')

        self.progressbar = ttk.Progressbar(prog_frame, mode='determinate',
                                           length=660)
        self.progressbar.pack(fill='x', pady=(2, 0))

        # Log ao vivo
        log_frame = tk.Frame(self.root, padx=20, pady=4)
        log_frame.pack(fill='both', expand=True)

        self.log = scrolledtext.ScrolledText(
            log_frame, height=16, state='disabled',
            font=('Consolas', 9), wrap='word',
            bg='#1e1e1e', fg='#d4d4d4',
            insertbackground='white'
        )
        self.log.pack(fill='both', expand=True)

        # Tags de cor no log
        self.log.tag_config('info',  foreground='#9cdcfe')
        self.log.tag_config('novo',  foreground='#4ec9b0')
        self.log.tag_config('aviso', foreground='#dcdcaa')
        self.log.tag_config('erro',  foreground='#f44747')

        # Botões
        btn_frame = tk.Frame(self.root, padx=20, pady=10)
        btn_frame.pack(fill='x')

        self.btn_iniciar = tk.Button(
            btn_frame, text='Iniciar busca',
            command=self._iniciar,
            bg=COR_VERDE, fg='white',
            font=('Segoe UI', 10, 'bold'),
            padx=20, pady=7, cursor='hand2',
            relief='flat', activebackground=COR_VERDE_ESC
        )
        self.btn_iniciar.pack(side='left')

        self.btn_parar = tk.Button(
            btn_frame, text='Parar',
            command=self._parar,
            bg='#6b7280', fg='white',
            font=('Segoe UI', 10),
            padx=16, pady=7, cursor='hand2',
            relief='flat', state='disabled'
        )
        self.btn_parar.pack(side='left', padx=(8, 0))

        self.var_status = tk.StringVar(
            value='Clique em "Iniciar busca" para consultar novos proventos.')
        tk.Label(btn_frame, textvariable=self.var_status,
                 font=('Segoe UI', 9), fg='gray'
                 ).pack(side='left', padx=(16, 0))

    def _card(self, parent, titulo, valor, col):
        frame = tk.Frame(parent, bg='#f5f5f0', padx=8, pady=6)
        frame.grid(row=0, column=col, sticky='nsew', padx=(0, 8))
        tk.Label(frame, text=titulo, bg='#f5f5f0',
                 font=('Segoe UI', 8), fg='gray').pack(anchor='w')
        var = tk.StringVar(value=valor)
        tk.Label(frame, textvariable=var, bg='#f5f5f0',
                 font=('Segoe UI', 11, 'bold')).pack(anchor='w')
        return var

    # ------------------------------------------------------------------
    # Log
    # ------------------------------------------------------------------
    def _log(self, msg, tag='info'):
        """Thread-safe: agenda a escrita no log na thread principal."""
        self.root.after(0, self._escrever_log, msg, tag)

    def _escrever_log(self, msg, tag):
        hora = datetime.now().strftime('%H:%M:%S')
        self.log.config(state='normal')
        self.log.insert('end', f'[{hora}] {msg}\n', tag)
        self.log.see('end')
        self.log.config(state='disabled')

    # ------------------------------------------------------------------
    # Progresso
    # ------------------------------------------------------------------
    def _progresso(self, i, total):
        self.root.after(0, self._atualizar_progresso, i, total)

    def _atualizar_progresso(self, i, total):
        pct = int(i / total * 100) if total else 0
        self.progressbar['value'] = pct
        self.var_prog_label.set(f'Consultando ativo {i} de {total}  ({pct}%)')

    # ------------------------------------------------------------------
    # Controle
    # ------------------------------------------------------------------
    def _iniciar(self):
        if self.em_execucao:
            return
        self.em_execucao = True
        self.parar_evt.clear()

        # Limpa estado anterior
        self.log.config(state='normal')
        self.log.delete('1.0', 'end')
        self.log.config(state='disabled')
        self.progressbar['value'] = 0
        self.var_prog_label.set('Iniciando...')
        for card in (self.card_tickers, self.card_inseridos,
                     self.card_ignorados, self.card_erros):
            card.set('—')

        self.btn_iniciar.config(state='disabled')
        self.btn_parar.config(state='normal')
        self.var_status.set('Buscando proventos...')

        threading.Thread(target=self._executar, daemon=True).start()

    def _parar(self):
        self.parar_evt.set()
        self.var_status.set('Interrompendo após o ativo atual...')
        self.btn_parar.config(state='disabled')

    def _executar(self):
        try:
            stats = rastrear_e_inserir_dividendos(
                log_cb       = self._log,
                progresso_cb = self._progresso,
                parar_evento = self.parar_evt,
            )
            self.root.after(0, self._finalizar, stats)
        except Exception as e:
            self.root.after(0, self._finalizar_erro, str(e))

    def _finalizar(self, stats):
        self.em_execucao = False
        self.btn_iniciar.config(state='normal')
        self.btn_parar.config(state='disabled')
        self.progressbar['value'] = 100
        self.var_prog_label.set('Concluído.')

        self.card_tickers.set(str(stats['tickers']))
        self.card_inseridos.set(str(stats['inseridos']))
        self.card_ignorados.set(str(stats['ignorados']))
        self.card_erros.set(str(stats['erros']))

        if stats['inseridos'] > 0:
            self.var_status.set(
                f"{stats['inseridos']} novo(s) provento(s) registrado(s) com sucesso.")
        else:
            self.var_status.set('Tudo atualizado — nenhum provento novo encontrado.')

    def _finalizar_erro(self, mensagem):
        self.em_execucao = False
        self.btn_iniciar.config(state='normal')
        self.btn_parar.config(state='disabled')
        self.var_status.set('Erro durante a execução.')
        messagebox.showerror('Erro', mensagem)

    # ------------------------------------------------------------------
    def _centralizar(self, largura, altura):
        self.root.update_idletasks()
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        self.root.geometry(
            f'{largura}x{altura}+{(sw-largura)//2}+{(sh-altura)//2}')


def main():
    root = tk.Tk()
    CacaDividendosApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()
