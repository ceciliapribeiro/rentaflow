import pdfplumber
import re
import os
import sys
import shutil
import threading
import openpyxl
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
from datetime import datetime
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config_loader import carregar_config, fazer_backup

_cfg          = carregar_config()
ARQUIVO_EXCEL = _cfg['arquivo_excel']
PASTA_NOTAS   = _cfg['pasta_notas']
SENHA_CPF     = _cfg['senha_pdf']

ABA_OPERACOES = 'OPERAÇÕES'

COR_VERDE     = '#1a6b45'
COR_VERDE_ESC = '#145535'
COR_VERMELHO  = '#991b1b'


# ══════════════════════════════════════════════════════════════════════
# LÓGICA DE NEGÓCIO
# ══════════════════════════════════════════════════════════════════════

def _carregar_indice_short_names():
    """
    Lê a coluna 'Short Name' (col I = índice 9) da aba Dados B3 e
    constrói um dicionário NOME_UPPER → TICKER para resolução local.
    Também indexa variantes sem sufixos de tipo/evento corporativo.
    """
    # Mesma regex de sufixos usada em _resolver_ticker
    sufixos = (
        r'\s+(?:'
        r'FRACIONARIO|VISTA|S\.A\.|S/A|UNIT|UNT|DRN|BDR|'
        r'EDJ|EJB|EDR|EDB|EJD|ATZ|ON|PN|CI|NM|N[123]|ED|EJ|ER|EB|REC|PNA|PNB'
        r')\b'
    )

    indice = {}
    try:
        wb    = openpyxl.load_workbook(ARQUIVO_EXCEL, data_only=True)
        sheet = wb['Dados B3']
        for row in range(2, sheet.max_row + 1):
            ticker = sheet.cell(row=row, column=1).value   # col A
            short  = sheet.cell(row=row, column=9).value   # col I — Short Name
            if not ticker or not short:
                continue
            ticker = str(ticker).upper().strip()
            short  = str(short).upper().strip()
            indice[short] = ticker
            # Variante sem sufixos
            short_limpo = short
            for _ in range(5):
                novo = re.sub(sufixos, '', short_limpo).strip()
                novo = re.sub(r'\s{2,}', ' ', novo).strip()
                if novo == short_limpo:
                    break
                short_limpo = novo
            if short_limpo and short_limpo != short:
                indice.setdefault(short_limpo, ticker)
        wb.close()
    except Exception:
        pass
    return indice


def _resolver_ticker(nome_bruto, indice):
    """
    Tenta resolver o nome de pregão da nota no ticker usando o índice.
    Estratégia em 4 níveis:
      1. Match exato
      2. Match após remover sufixos de tipo/governança/evento corporativo da B3
      3. Ticker embutido no nome (ex: "FII TRX REAL TRXF11 CI ER")
      4. Match parcial — exige ao menos 2 palavras significativas em comum

    Sufixos da B3 tratados:
      Tipo de ativo : ON, PN, UNT, CI, BDR, DRN
      Governança   : NM, N1, N2, N3
      Eventos corp.: ED, EJ, ER, EB, ATZ, EDJ, EJB, EDR, EDB, EJD, REC, PNA, PNB
      Mercado      : FRACIONARIO, VISTA
      Outros       : S.A., S/A, UNIT, FRAC
    """
    # Regex que remove sufixos individualmente (ordem importa: mais longos primeiro)
    sufixos = (
        r'\s+(?:'
        r'FRACIONARIO|VISTA|'          # mercado
        r'S\.A\.|S/A|'                 # razão social
        r'UNIT|UNT|DRN|BDR|'          # tipo de ativo
        r'EDJ|EJB|EDR|EDB|EJD|'       # eventos compostos (3 letras — antes dos simples)
        r'ATZ|'                        # subscrição/atualização
        r'ON|PN|CI|'                   # tipo
        r'NM|N[123]|'                  # governança
        r'ED|EJ|ER|EB|'               # eventos simples
        r'REC|PNA|PNB'                 # outros
        r')\b'
    )

    nome = nome_bruto.upper().strip()

    # Nível 1 — exato
    if nome in indice:
        return indice[nome]

    # Nível 2 — remove todos os sufixos iterativamente até estabilizar
    nome_limpo = nome
    for _ in range(5):   # no máximo 5 passes para remover sufixos compostos
        novo = re.sub(sufixos, '', nome_limpo).strip()
        novo = re.sub(r'\s{2,}', ' ', novo).strip()
        if novo == nome_limpo:
            break
        nome_limpo = novo

    if nome_limpo and nome_limpo in indice:
        return indice[nome_limpo]

    # Nível 3 — ticker embutido no nome de pregão
    # Ex: "FII TRX REAL TRXF11 CI ER" → TRXF11
    m_embutido = re.search(r'\b([A-Z]{4}\d{1,2}F?)\b', nome)
    if m_embutido:
        ticker_embutido = m_embutido.group(1)
        if ticker_embutido.endswith('F') and len(ticker_embutido) > 4:
            ticker_embutido = ticker_embutido[:-1]
        return ticker_embutido

    # Nível 4 — match parcial com ao menos 2 palavras significativas em comum
    palavras_nota = set(w for w in nome_limpo.split() if len(w) > 2)

    melhor_ticker = None
    melhor_score  = 0

    for chave, ticker in indice.items():
        chave_limpa = chave
        for _ in range(5):
            novo = re.sub(sufixos, '', chave_limpa).strip()
            novo = re.sub(r'\s{2,}', ' ', novo).strip()
            if novo == chave_limpa:
                break
            chave_limpa = novo
        palavras_chave = set(w for w in chave_limpa.split() if len(w) > 2)

        intersecao = palavras_nota & palavras_chave
        if len(intersecao) >= 2 and len(intersecao) > melhor_score:
            melhor_score  = len(intersecao)
            melhor_ticker = ticker

    return melhor_ticker


def extrair_dados_pdf(caminho_pdf, log_cb, indice_short_names=None):
    """
    Extrai operações do PDF.
    indice_short_names: dict {SHORT_NAME → TICKER} carregado uma vez
    pelo chamador para evitar releitura da planilha a cada PDF.
    """
    operacoes      = []
    nome_arquivo   = os.path.basename(caminho_pdf)
    nao_resolvidos = []

    try:
        with pdfplumber.open(caminho_pdf, password=SENHA_CPF) as pdf:
            texto_completo = ''
            for pagina in pdf.pages:
                texto_completo += pagina.extract_text() + '\n'

            # ── Data do pregão ───────────────────────────────────────
            match_data = re.search(
                r'Data pregão\s*[:\n]?\s*(\d{2}/\d{2}/\d{4})',
                texto_completo, re.IGNORECASE)
            if match_data:
                data_operacao = match_data.group(1)
            else:
                match_data_arq = re.search(
                    r'(\d{2})[-_](\d{2})[-_](\d{4})', nome_arquivo)
                if match_data_arq:
                    data_operacao = (f"{match_data_arq.group(1)}/"
                                     f"{match_data_arq.group(2)}/"
                                     f"{match_data_arq.group(3)}")
                else:
                    data_operacao = datetime.today().strftime('%d/%m/%Y')

            # ── Padrão único: usa D/C no final como fonte de verdade ──
            # D = Débito = COMPRA  |  C = Crédito = VENDA
            # Formato: "1-BOVESPA ... NOME/TICKER ... QTDE  PREÇO  TOTAL  D|C"
            # O @, @# etc. no campo Obs.(*) são ignorados — não são separadores reais.
            #
            # Grupos: (nome_ou_ticker)  (qtde)  (preco_unit)  (total)  (D|C)
            padrao = (
                r'BOVESPA\s+[CV]\s+'
                r'(?:VISTA|FRACIONARIO|FRAC)?\s*'
                r'(.+?)\s+'               # nome ou ticker (ganancioso — pega tudo até os números)
                r'@[^0-9\-]*'             # marcador Obs.(*): @, @#, @*, etc. — descartado
                r'(-?[\d\.]+)\s+'         # quantidade
                r'([\d,\.]+)\s+'          # preço unitário
                r'[\d,\.]+\s+'            # valor total — descartado
                r'([DC])\b'               # D=compra  C=venda  ← fonte de verdade
            )

            for linha in texto_completo.split('\n'):
                if 'BOVESPA' not in linha:
                    continue

                m = re.search(padrao, linha)
                if not m:
                    continue

                nome_bruto = m.group(1).strip()
                qtde_str   = m.group(2).replace('.', '').replace('-', '')
                qtde       = float(qtde_str) if qtde_str else 0.0
                preco      = float(m.group(3).replace('.', '').replace(',', '.'))
                dc         = m.group(4)
                tipo_cv    = 'COMPRA' if dc == 'D' else 'VENDA'

                if qtde <= 0:
                    continue

                # Verifica se o nome é um ticker direto (ex: MXRF11)
                m_ticker = re.match(r'^([A-Z]{4}\d{1,2})(F?)$', nome_bruto.split()[0])
                if m_ticker:
                    ticker = m_ticker.group(1)
                    operacoes.append({
                        'data': data_operacao, 'operacao': tipo_cv,
                        'ticker': ticker, 'qtde': qtde, 'preco': preco,
                        'origem': nome_arquivo,
                    })
                    continue

                # Caso contrário resolve pelo nome de pregão
                # Remove sufixos FII embutidos no nome (ex: "FII TRX REAL TRXF11 CI ER")
                ticker = _resolver_ticker(nome_bruto, indice_short_names or {})

                if ticker:
                    operacoes.append({
                        'data': data_operacao, 'operacao': tipo_cv,
                        'ticker': ticker, 'qtde': qtde, 'preco': preco,
                        'origem': nome_arquivo,
                    })
                    log_cb(f'  → "{nome_bruto}" → {ticker}', 'info')
                else:
                    nao_resolvidos.append({
                        'nome': nome_bruto, 'tipo': tipo_cv,
                        'qtde': qtde, 'preco': preco, 'data': data_operacao,
                    })

    except Exception as e:
        log_cb(f'[ERRO] {nome_arquivo}: {e}', 'erro')
        return []

    if nao_resolvidos:
        log_cb(f'[AVISO] {len(nao_resolvidos)} operação(ões) sem ticker — '
               f'adicione o Short Name na aba Dados B3 e rode Atualizar Cotações:', 'aviso')
        for nr in nao_resolvidos:
            log_cb(f'  → {nr["tipo"]} | "{nr["nome"]}" | '
                   f'{nr["qtde"]:.0f}x R$ {nr["preco"]:.2f} ({nr["data"]})', 'aviso')

    return operacoes




def ler_operacoes_existentes(sheet):
    existentes = set()
    for row in range(2, sheet.max_row + 1):
        data_raw = sheet.cell(row=row, column=1).value
        ticker   = sheet.cell(row=row, column=2).value
        qtde     = sheet.cell(row=row, column=3).value
        operacao = sheet.cell(row=row, column=8).value
        if not ticker or not data_raw or qtde is None or not operacao:
            continue
        try:
            if isinstance(data_raw, datetime):
                data_str = data_raw.strftime('%d/%m/%Y')
            else:
                data_str = str(data_raw).split(' ')[0]
                if '-' in data_str and len(data_str.split('-')[0]) == 4:
                    p = data_str.split('-')
                    data_str = f"{p[2]}/{p[1]}/{p[0]}"
            existentes.add(
                f"{data_str}_{str(ticker).upper().strip()}"
                f"_{float(qtde):.0f}_{str(operacao).upper().strip()}"
            )
        except:
            pass
    return existentes


def calcular_preco_medio(sheet, ticker, data_operacao):
    """
    Reconstrói o preço médio de custo do ticker a partir do histórico
    de operações já registradas na planilha, considerando apenas
    operações com data <= data_operacao.

    Retorna o preço médio atual, ou 0.0 se não houver posição.
    """
    # Normaliza a data de referência
    if isinstance(data_operacao, str):
        try:
            ref = datetime.strptime(data_operacao, '%d/%m/%Y')
        except ValueError:
            try:
                ref = datetime.strptime(data_operacao, '%Y-%m-%d')
            except ValueError:
                ref = datetime.max
    else:
        ref = data_operacao or datetime.max

    qtde_acum  = 0.0
    custo_acum = 0.0

    for row in range(2, sheet.max_row + 1):
        t = sheet.cell(row=row, column=2).value  # Ticker (col B)
        if not t or str(t).upper().strip() != ticker.upper().strip():
            continue

        data_raw = sheet.cell(row=row, column=1).value  # Data (col A)
        if not data_raw:
            continue

        # Converte a data da linha
        if isinstance(data_raw, datetime):
            data_linha = data_raw.replace(hour=0, minute=0, second=0, microsecond=0)
        else:
            s = str(data_raw).split(' ')[0]
            try:
                data_linha = datetime.strptime(s, '%Y-%m-%d')
            except ValueError:
                try:
                    data_linha = datetime.strptime(s, '%d/%m/%Y')
                except ValueError:
                    continue

        if data_linha > ref:
            continue

        tipo = str(sheet.cell(row=row, column=8).value or '').upper()  # Operação (col H)
        try:
            qtde  = abs(float(sheet.cell(row=row, column=3).value or 0))  # Qtde (col C)
            preco = float(sheet.cell(row=row, column=4).value or 0)       # Preço unit. (col D)
        except (TypeError, ValueError):
            continue

        if 'COMPRA' in tipo:
            custo_acum += qtde * preco
            qtde_acum  += qtde
        elif 'VENDA' in tipo:
            if qtde_acum > 0:
                preco_medio = custo_acum / qtde_acum
                custo_acum -= qtde * preco_medio
                qtde_acum  -= qtde
                custo_acum  = max(custo_acum, 0.0)
                qtde_acum   = max(qtde_acum, 0.0)

    if qtde_acum > 0:
        return round(custo_acum / qtde_acum, 4)
    return 0.0


def gravar_operacoes(lista_operacoes, log_cb):
    """Grava apenas as operações da lista (já filtradas/confirmadas pelo usuário)."""
    fazer_backup(ARQUIVO_EXCEL, log_cb)

    wb    = openpyxl.load_workbook(ARQUIVO_EXCEL)
    sheet = wb[ABA_OPERACOES]
    existentes = ler_operacoes_existentes(sheet)

    linha_insercao = 2
    for row in range(2, sheet.max_row + 10):
        if sheet.cell(row=row, column=2).value is None:
            linha_insercao = row
            break

    inseridas  = 0
    duplicadas = 0

    for op in lista_operacoes:
        chave = (f"{op['data']}_{op['ticker'].upper()}"
                 f"_{float(op['qtde']):.0f}_{op['operacao'].upper()}")
        if chave in existentes:
            log_cb(f'[DUPLICATA] {op["operacao"]} {op["ticker"]} '
                   f'{op["qtde"]:.0f} cotas {op["data"]}', 'aviso')
            duplicadas += 1
            continue

        # Col F (6) — Qtde executada: negativo para vendas, positivo para compras
        qtde_exec = op['qtde'] if op['operacao'] == 'COMPRA' else -op['qtde']

        if op['operacao'] == 'VENDA':
            # Col D = preço médio de custo (calculado a partir do histórico)
            # Col I = preço real de venda (extraído da nota)
            preco_medio = calcular_preco_medio(sheet, op['ticker'], op['data'])
            preco_col_d = preco_medio if preco_medio > 0 else op['preco']
            preco_col_i = op['preco']   # preço real de execução da nota

            if preco_medio <= 0:
                log_cb(f'[AVISO] {op["ticker"]}: preço médio não encontrado — '
                       f'usando preço da nota na col D. Verifique as compras anteriores.',
                       'aviso')
        else:
            # COMPRA: col D = preço de execução, col I não se aplica
            preco_col_d = op['preco']
            preco_col_i = None

        sheet.cell(row=linha_insercao, column=1).value = op['data']
        sheet.cell(row=linha_insercao, column=2).value = op['ticker']
        sheet.cell(row=linha_insercao, column=3).value = op['qtde']
        sheet.cell(row=linha_insercao, column=4).value = preco_col_d
        sheet.cell(row=linha_insercao, column=6).value = qtde_exec
        if preco_col_i is not None:
            sheet.cell(row=linha_insercao, column=9).value = preco_col_i
        sheet.cell(row=linha_insercao, column=8).value = op['operacao']

        if op['operacao'] == 'VENDA':
            log_cb(f'[NOVO] VENDA   {op["ticker"]:<8} {op["qtde"]:.0f}x  '
                   f'PM: R$ {preco_col_d:.4f}  VL real: R$ {preco_col_i:.2f}  ({op["data"]})',
                   'novo')
        else:
            log_cb(f'[NOVO] COMPRA  {op["ticker"]:<8} {op["qtde"]:.0f}x  '
                   f'R$ {preco_col_d:.2f}  ({op["data"]})', 'novo')

        existentes.add(chave)
        linha_insercao += 1
        inseridas += 1

    if inseridas > 0:
        wb.save(ARQUIVO_EXCEL)
    wb.close()

    return inseridas, duplicadas


# ══════════════════════════════════════════════════════════════════════
# INTERFACE GRÁFICA
# ══════════════════════════════════════════════════════════════════════

class LeitorNotasApp:
    def __init__(self, root):
        self.root            = root
        self.operacoes_lidas = []   # resultado bruto da extração
        self.em_execucao     = False

        self.root.title('Importar Notas de Corretagem')
        self.root.resizable(False, False)
        self._centralizar(760, 620)
        self._construir_ui()

    # ------------------------------------------------------------------
    # Layout
    # ------------------------------------------------------------------
    def _construir_ui(self):
        # Cabeçalho
        header = tk.Frame(self.root, bg=COR_VERDE, height=60)
        header.pack(fill='x')
        header.pack_propagate(False)
        tk.Label(header, text='  Importar Notas de Corretagem (PDF)',
                 bg=COR_VERDE, fg='white',
                 font=('Segoe UI', 13, 'bold'), anchor='w'
                 ).pack(fill='both', expand=True, padx=4)

        # Seleção de pasta
        pasta_frame = tk.Frame(self.root, padx=20, pady=10)
        pasta_frame.pack(fill='x')

        tk.Label(pasta_frame, text='Pasta das notas:',
                 font=('Segoe UI', 9)).grid(row=0, column=0, sticky='w')

        self.var_pasta = tk.StringVar(value=PASTA_NOTAS)
        tk.Entry(pasta_frame, textvariable=self.var_pasta,
                 font=('Segoe UI', 9), width=52, state='readonly'
                 ).grid(row=0, column=1, padx=(8, 0), sticky='ew')

        tk.Button(pasta_frame, text='Alterar...',
                  command=self._escolher_pasta,
                  font=('Segoe UI', 9), cursor='hand2'
                  ).grid(row=0, column=2, padx=(8, 0))

        self.var_pdfs = tk.StringVar(value='')
        tk.Label(pasta_frame, textvariable=self.var_pdfs,
                 font=('Segoe UI', 8), fg='gray'
                 ).grid(row=1, column=0, columnspan=3, sticky='w', pady=(4, 0))

        pasta_frame.columnconfigure(1, weight=1)
        self._contar_pdfs()

        # Notebook: Preview | Log
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True, padx=20, pady=(4, 0))

        # Aba 1: Tabela de preview
        aba_preview = tk.Frame(self.notebook)
        self.notebook.add(aba_preview, text='  Operações extraídas  ')

        cols = ('origem', 'data', 'operacao', 'ticker', 'qtde', 'preco')
        self.tree = ttk.Treeview(aba_preview, columns=cols,
                                 show='headings', height=14)
        defs = [
            ('origem',   'Arquivo PDF',  180, 'w'),
            ('data',     'Data',          80, 'center'),
            ('operacao', 'Tipo',          70, 'center'),
            ('ticker',   'Ativo',         70, 'center'),
            ('qtde',     'Qtde',          60, 'center'),
            ('preco',    'Preço unit.',   90, 'e'),
        ]
        for col, titulo, larg, ancora in defs:
            self.tree.heading(col, text=titulo)
            self.tree.column(col, width=larg, anchor=ancora, stretch=False)

        self.tree.tag_configure('compra', foreground='#166534')
        self.tree.tag_configure('venda',  foreground='#991b1b')
        self.tree.tag_configure('par',    background='#f9f9f6')
        self.tree.tag_configure('impar',  background='#ffffff')

        scroll_t = ttk.Scrollbar(aba_preview, orient='vertical',
                                 command=self.tree.yview)
        self.tree.configure(yscrollcommand=scroll_t.set)
        self.tree.pack(side='left', fill='both', expand=True)
        scroll_t.pack(side='right', fill='y')

        # Aba 2: Log
        aba_log = tk.Frame(self.notebook)
        self.notebook.add(aba_log, text='  Log  ')

        self.log = scrolledtext.ScrolledText(
            aba_log, height=14, state='disabled',
            font=('Consolas', 9), wrap='word',
            bg='#1e1e1e', fg='#d4d4d4',
        )
        self.log.pack(fill='both', expand=True)
        self.log.tag_config('info',  foreground='#9cdcfe')
        self.log.tag_config('novo',  foreground='#4ec9b0')
        self.log.tag_config('aviso', foreground='#dcdcaa')
        self.log.tag_config('erro',  foreground='#f44747')

        # Rodapé
        btn_frame = tk.Frame(self.root, padx=20, pady=10)
        btn_frame.pack(fill='x')

        self.btn_ler = tk.Button(
            btn_frame, text='1. Ler PDFs',
            command=self._ler_pdfs,
            bg='#374151', fg='white',
            font=('Segoe UI', 10, 'bold'),
            padx=18, pady=7, cursor='hand2', relief='flat'
        )
        self.btn_ler.pack(side='left')

        self.btn_importar = tk.Button(
            btn_frame, text='2. Confirmar e importar',
            command=self._confirmar_importar,
            bg=COR_VERDE, fg='white',
            font=('Segoe UI', 10, 'bold'),
            padx=18, pady=7, cursor='hand2',
            relief='flat', activebackground=COR_VERDE_ESC,
            state='disabled'
        )
        self.btn_importar.pack(side='left', padx=(8, 0))

        self.var_status = tk.StringVar(
            value='Clique em "1. Ler PDFs" para extrair as operações.')
        tk.Label(btn_frame, textvariable=self.var_status,
                 font=('Segoe UI', 9), fg='gray', wraplength=340
                 ).pack(side='left', padx=(16, 0))

    # ------------------------------------------------------------------
    # Utilitários
    # ------------------------------------------------------------------
    def _contar_pdfs(self):
        pasta = self.var_pasta.get()
        if os.path.isdir(pasta):
            n = len([f for f in os.listdir(pasta) if f.lower().endswith('.pdf')])
            self.var_pdfs.set(f'{n} arquivo(s) PDF encontrado(s) na pasta.')
        else:
            self.var_pdfs.set('Pasta não encontrada.')

    def _escolher_pasta(self):
        nova = filedialog.askdirectory(title='Selecione a pasta das notas de corretagem')
        if nova:
            self.var_pasta.set(nova)
            self._contar_pdfs()

    def _log(self, msg, tag='info'):
        self.root.after(0, self._escrever_log, msg, tag)

    def _escrever_log(self, msg, tag):
        hora = datetime.now().strftime('%H:%M:%S')
        self.log.config(state='normal')
        self.log.insert('end', f'[{hora}] {msg}\n', tag)
        self.log.see('end')
        self.log.config(state='disabled')

    def _centralizar(self, largura, altura):
        self.root.update_idletasks()
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        self.root.geometry(
            f'{largura}x{altura}+{(sw-largura)//2}+{(sh-altura)//2}')

    # ------------------------------------------------------------------
    # Passo 1 — Ler PDFs
    # ------------------------------------------------------------------
    def _ler_pdfs(self):
        pasta = self.var_pasta.get()
        if not os.path.isdir(pasta):
            messagebox.showerror('Pasta não encontrada',
                                 f'A pasta abaixo não existe:\n{pasta}\n\n'
                                 'Clique em "Alterar..." para escolher a pasta correta.')
            return

        pdfs = [f for f in os.listdir(pasta) if f.lower().endswith('.pdf')]
        if not pdfs:
            messagebox.showwarning('Nenhum PDF',
                                   f'Nenhum arquivo PDF encontrado em:\n{pasta}')
            return

        # Limpa estado anterior
        for row in self.tree.get_children():
            self.tree.delete(row)
        self.operacoes_lidas = []
        self.btn_importar.config(state='disabled')
        self.btn_ler.config(state='disabled', text='Lendo...')
        self.var_status.set(f'Lendo {len(pdfs)} PDF(s)...')
        self.notebook.select(1)   # vai para aba Log

        threading.Thread(
            target=self._processar_pdfs, args=(pasta, pdfs),
            daemon=True).start()

    def _processar_pdfs(self, pasta, pdfs):
        # Carrega o índice Short Name → Ticker uma única vez para todos os PDFs
        indice = _carregar_indice_short_names()
        if indice:
            self._log(f'Índice carregado: {len(indice)} nomes de pregão mapeados.', 'info')
        else:
            self._log('[AVISO] Short Name não encontrado na planilha. '
                      'Execute "Atualizar Cotações" para popular a coluna I da aba Dados B3.', 'aviso')

        todas = []
        for arquivo in pdfs:
            caminho = os.path.join(pasta, arquivo)
            self._log(f'Lendo: {arquivo}', 'info')
            ops = extrair_dados_pdf(caminho, self._log, indice)
            if ops:
                todas.extend(ops)
                self._log(f'  → {len(ops)} operação(ões) extraída(s).', 'info')
            else:
                self._log(f'  → Nenhuma operação identificada.', 'aviso')

        self.root.after(0, self._exibir_preview, todas)

    def _exibir_preview(self, operacoes):
        self.operacoes_lidas = operacoes
        self.btn_ler.config(state='normal', text='1. Ler PDFs')

        # Popula tabela
        for i, op in enumerate(operacoes):
            tags = []
            tags.append('compra' if op['operacao'] == 'COMPRA' else 'venda')
            tags.append('par' if i % 2 == 0 else 'impar')
            self.tree.insert('', 'end', tags=tags, values=(
                op['origem'],
                op['data'],
                op['operacao'],
                op['ticker'],
                f"{op['qtde']:.0f}",
                f"R$ {op['preco']:.2f}",
            ))

        self.notebook.select(0)   # vai para aba Preview

        if not operacoes:
            self.var_status.set('Nenhuma operação extraída dos PDFs.')
        else:
            self.btn_importar.config(state='normal')
            self.var_status.set(
                f'{len(operacoes)} operação(ões) extraída(s). '
                f'Revise a tabela e clique em "2. Confirmar e importar".')

    # ------------------------------------------------------------------
    # Passo 2 — Confirmar e importar
    # ------------------------------------------------------------------
    def _confirmar_importar(self):
        if not self.operacoes_lidas:
            return

        n = len(self.operacoes_lidas)
        confirmado = messagebox.askyesno(
            'Confirmar importação',
            f'Você está prestes a importar {n} operação(ões) para a planilha.\n\n'
            f'Arquivo: {os.path.basename(ARQUIVO_EXCEL)}\n\n'
            f'Deseja continuar?'
        )
        if not confirmado:
            return

        self.btn_importar.config(state='disabled', text='Importando...')
        self.btn_ler.config(state='disabled')
        self.notebook.select(1)

        threading.Thread(
            target=self._gravar, daemon=True).start()

    def _gravar(self):
        try:
            inseridas, duplicadas = gravar_operacoes(self.operacoes_lidas, self._log)
            self.root.after(0, self._finalizar, inseridas, duplicadas)
        except Exception as e:
            self.root.after(0, self._finalizar_erro, str(e))

    def _finalizar(self, inseridas, duplicadas):
        self.btn_ler.config(state='normal')
        self.btn_importar.config(state='normal', text='2. Confirmar e importar')

        partes = [f'{inseridas} operação(ões) importada(s).']
        if duplicadas:
            partes.append(f'{duplicadas} duplicata(s) ignorada(s).')
        self.var_status.set('  '.join(partes))

        if inseridas > 0:
            messagebox.showinfo(
                'Importação concluída',
                f'{inseridas} operação(ões) gravada(s) na planilha.\n'
                + (f'{duplicadas} duplicata(s) ignorada(s).' if duplicadas else '')
            )

    def _finalizar_erro(self, mensagem):
        self.btn_ler.config(state='normal')
        self.btn_importar.config(state='normal', text='2. Confirmar e importar')
        self.var_status.set('Erro durante a importação.')
        messagebox.showerror('Erro', mensagem)


def main():
    root = tk.Tk()
    LeitorNotasApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()