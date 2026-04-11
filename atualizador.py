import openpyxl
import yfinance as yf
import requests
from bs4 import BeautifulSoup
import shutil
import time
import re
import os
import sys
import threading
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
from datetime import datetime
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config_loader import carregar_config, fazer_backup

_cfg          = carregar_config()
ARQUIVO_EXCEL = _cfg['arquivo_excel']

# ==========================================
# CONFIGURAÇÃO DE AMBIENTE
# ==========================================
ABA_DADOS    = 'Dados B3'
ABA_CARTEIRA = 'CARTEIRA'

COL_TICKER     = 1
COL_TIPO       = 2
COL_PRECO      = 4
COL_CNPJ       = 5
COL_DY         = 7
COL_PVP        = 8  # H
COL_SEG        = 6  # F — Segmento
COL_SHORT_NAME = 9  # I — Short Name (nome de pregão B3)

SALVAR_A_CADA = 30
FIAGROS       = {'AAZQ11', 'VGIA11', 'SNAG11', 'RURA11', 'KNCA11'}

COR_VERDE     = '#1a6b45'
COR_VERDE_ESC = '#145535'
COR_AMARELO   = '#b45309'
# ==========================================


# ══════════════════════════════════════════════════════════════════════
# LÓGICA DE NEGÓCIO
# ══════════════════════════════════════════════════════════════════════

def carregar_ativos_da_carteira():
    try:
        wb    = openpyxl.load_workbook(ARQUIVO_EXCEL, data_only=True)
        sheet = wb[ABA_CARTEIRA]
        ativos = set()
        for row in range(3, sheet.max_row + 1):
            t = sheet.cell(row=row, column=12).value
            if t and isinstance(t, str):
                ticker = t.strip().upper()
                if ticker and ticker not in ('ATIVO', 'TOTAL', 'TOTAL GERAL'):
                    ativos.add(ticker)
        wb.close()
        return ativos
    except:
        return set()


def _extrair_indicador_si(soup, html_texto, siglas):
    """
    Tenta extrair o valor numérico de um indicador do Status Invest por múltiplas
    estratégias. Retorna string no formato '1,05' ou None se não encontrar.

    Estratégias (em ordem):
      1. div[title] contendo qualquer das siglas
      2. div[data-title] contendo qualquer das siglas
      3. Qualquer <div> cujo texto de cabeçalho contenha a sigla
      4. Regex direto no HTML — última linha de defesa
    """
    INVALIDOS = {'-%', '-', '', 'N/A', 'n/a'}

    def _limpar(txt):
        v = txt.strip()
        return v if v not in INVALIDOS else None

    # Estratégia 1 — atributo title
    for sigla in siglas:
        bloco = soup.find('div', title=lambda t, s=sigla: t and s in str(t))
        if bloco:
            strong = bloco.find('strong', class_='value')
            if strong:
                v = _limpar(strong.get_text(strip=True))
                if v:
                    return v

    # Estratégia 2 — atributo data-title
    for sigla in siglas:
        bloco = soup.select_one(f'div[data-title*="{sigla}"]')
        if bloco:
            strong = bloco.find('strong', class_='value')
            if strong:
                v = _limpar(strong.get_text(strip=True))
                if v:
                    return v

    # Estratégia 3 — percorre todos os divs em busca do label
    for sigla in siglas:
        for div in soup.find_all('div', class_=True):
            label = div.find(['span', 'h3', 'p'], string=re.compile(
                rf'^\s*{re.escape(sigla)}\s*$', re.I))
            if label:
                strong = div.find('strong', class_='value')
                if not strong:
                    strong = div.find('strong')
                if strong:
                    v = _limpar(strong.get_text(strip=True))
                    if v:
                        return v

    # Estratégia 4 — regex direto no HTML bruto
    for sigla in siglas:
        # Padrão: "P/VP":"1,05"  ou  P/VP</span>...<strong>1,05</strong>
        pat = re.search(
            rf'["\']?{re.escape(sigla)}["\']?\s*[":>][^<]{{0,60}}'
            rf'(\d{{1,4}}[,\.]\d{{1,4}})',
            html_texto, re.I | re.S
        )
        if pat:
            v = pat.group(1).strip()
            if v not in INVALIDOS:
                return v

    return None


def obter_pvp_yfinance(ticker):
    """
    Fallback de P/VP via yfinance (campo priceToBook).
    Retorna string '1,05' ou None.
    """
    try:
        info = yf.Ticker(f"{ticker}.SA").info
        ptb  = info.get('priceToBook')
        if ptb and 0 < float(ptb) < 100:
            return f"{float(ptb):.2f}".replace('.', ',')
    except Exception:
        pass
    return None


def obter_dados_status_invest(ticker, tipo, sessao, log_cb):
    headers   = {
        'User-Agent': (
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
            'AppleWebKit/537.36 (KHTML, like Gecko) '
            'Chrome/124.0.0.0 Safari/537.36'
        ),
        'Accept-Language': 'pt-BR,pt;q=0.9',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    }
    eh_fii    = str(tipo).strip().upper() == 'FII' or ticker.upper() in FIAGROS
    path      = 'fundos-imobiliarios' if eh_fii else 'acoes'
    url       = f"https://statusinvest.com.br/{path}/{ticker.lower()}"
    resultado = {'dy': '0,00', 'cnpj': '', 'pvp': '0,00', 'segmento': ''}

    for tentativa in range(3):
        try:
            req = sessao.get(url, headers=headers, timeout=12)

            if req.status_code in (429, 503):
                espera = 15 * (tentativa + 1)
                log_cb(f'Rate limit ({req.status_code}) em {ticker} — aguardando {espera}s...', 'aviso')
                time.sleep(espera)
                continue

            if req.status_code == 200:
                soup      = BeautifulSoup(req.text, 'html.parser')
                html_text = req.text

                # ── DY ────────────────────────────────────────────────
                dy_val = _extrair_indicador_si(
                    soup, html_text,
                    ['Dividend Yield', 'DY', 'D.Y.']
                )
                if dy_val:
                    resultado['dy'] = dy_val

                # ── P/VP ──────────────────────────────────────────────
                # FIIs: Status Invest chama de "P/VP"; ações de "P/VPA" ou "P/VP"
                siglas_pvp = ['P/VP', 'P/VPA', 'Preço/VP'] if not eh_fii else ['P/VP']
                pvp_val = _extrair_indicador_si(soup, html_text, siglas_pvp)
                if pvp_val:
                    resultado['pvp'] = pvp_val
                else:
                    # Fallback: yfinance priceToBook
                    pvp_yf = obter_pvp_yfinance(ticker)
                    if pvp_yf:
                        resultado['pvp'] = pvp_yf
                        log_cb(f'  P/VP de {ticker} via yfinance: {pvp_yf}', 'info')

                # ── CNPJ ──────────────────────────────────────────────
                m = re.search(r'\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}', html_text)
                if m:
                    resultado['cnpj'] = m.group(0)

                # ── Segmento ──────────────────────────────────────────
                seg_tag = soup.find('span', {'title': re.compile(r'Segmento|Setor|Categoria', re.I)})
                if not seg_tag:
                    seg_tag = soup.select_one('div.cell span.sub-value')
                if not seg_tag:
                    seg_meta = soup.find('meta', {'property': 'og:description'})
                    if seg_meta:
                        desc  = seg_meta.get('content', '')
                        m_seg = re.search(r'Segmento[:\s]+([^.\n,]+)', desc, re.I)
                        if m_seg:
                            resultado['segmento'] = m_seg.group(1).strip()
                else:
                    val_seg = seg_tag.text.strip()
                    if val_seg and val_seg not in ['-', '']:
                        resultado['segmento'] = val_seg

            break  # sucesso — sai do loop de tentativas

        except Exception:
            if tentativa < 2:
                time.sleep(3 * (tentativa + 1))

    return resultado


def obter_preco_yfinance(ticker):
    # Tickers BDR terminam em 34/35 e são negociados normalmente com .SA
    # Se não achar preço válido, tenta sem o sufixo como fallback
    for sufixo in ['.SA', '']:
        try:
            ativo = yf.Ticker(f"{ticker}{sufixo}")
            info  = ativo.fast_info
            preco = info.get('last_price') or info.get('regularMarketPrice')
            if preco and 0.01 < float(preco) < 1_000_000:
                return round(float(preco), 2)
            hist = ativo.history(period='5d')
            if not hist.empty:
                preco = float(hist['Close'].iloc[-1])
                if 0.01 < preco < 1_000_000:
                    return round(preco, 2)
        except:
            pass
    return None


def obter_dy_yfinance(ticker):
    """
    Fonte secundária de DY via yfinance.
    Usado quando o Status Invest retorna 0,00 ou vazio.
    Retorna string no formato '9,50' ou None se não encontrar.
    """
    try:
        info = yf.Ticker(f"{ticker}.SA").info
        dy   = info.get('dividendYield') or info.get('trailingAnnualDividendYield')
        if dy and dy > 0:
            return f"{dy * 100:.2f}".replace('.', ',')
    except Exception:
        pass
    return None


def obter_short_name_b3(ticker):
    """
    Busca o nome de pregão exato via API oficial da B3 (cotacao.b3.com.br).
    Retorna o campo 'desc' normalizado, que é a mesma fonte das notas de corretagem.
    Ex: BBDC4 → "BRADESCO PN N1", MXRF11 → "FII MAXI RENCI"
    Retorna string normalizada ou None.
    """
    try:
        url = f'https://cotacao.b3.com.br/mds/api/v1/instrumentQuotation/{ticker}'
        r   = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=8)
        if r.status_code == 200:
            data = r.json()
            trad = data.get('Trad', [])
            if trad:
                desc = trad[0].get('scty', {}).get('desc', '') or ''
                if desc:
                    # Normaliza espaços múltiplos: "BRADESCO    PN      N1" → "BRADESCO PN N1"
                    return ' '.join(desc.upper().split())
    except Exception:
        pass
    return None


def executar_atualizacao(log_cb, progresso_cb, parar_evento, apenas_carteira=False):
    """
    Lógica principal sem prints — usa callbacks para a UI.
    apenas_carteira=True: atualiza só os ativos da aba CARTEIRA (rápido).
    apenas_carteira=False: atualiza todo o catálogo Dados B3 (completo).
    Retorna dict com estatísticas finais.
    """
    stats = {'total': 0, 'atualizados': 0, 'sem_preco': [], 'erros': []}

    fazer_backup(ARQUIVO_EXCEL, log_cb)

    try:
        wb    = openpyxl.load_workbook(ARQUIVO_EXCEL)
        sheet = wb[ABA_DADOS]
    except Exception as e:
        raise RuntimeError(f"Erro ao abrir '{ABA_DADOS}': {e}")

    # Indexa ativos
    ativos_no_excel = {}
    for row in range(2, sheet.max_row + 1):
        t = sheet.cell(row=row, column=COL_TICKER).value
        if t:
            ativos_no_excel[str(t).strip().upper()] = row

    ativos_prioritarios = carregar_ativos_da_carteira()

    # Monta fila conforme o modo escolhido
    fila = []
    for ticker, linha in ativos_no_excel.items():
        eh_carteira = ticker in ativos_prioritarios
        # Modo carteira: inclui apenas ativos que estão na carteira
        if apenas_carteira and not eh_carteira:
            continue
        tipo = sheet.cell(row=linha, column=COL_TIPO).value
        fila.append({
            'ticker':     ticker,
            'linha':      linha,
            'tipo':       tipo,
            'prioridade': eh_carteira,
        })
    fila.sort(key=lambda x: x['prioridade'], reverse=True)

    total          = len(fila)
    stats['total'] = total
    modo_desc      = 'carteira' if apenas_carteira else 'catálogo completo'
    log_cb(f'Modo: {modo_desc}  |  {total} ativo(s) na fila  |  {len(ativos_prioritarios)} na carteira', 'info')
    log_cb('Iniciando atualização...', 'info')

    sessao     = requests.Session()
    inicio_ts  = time.time()

    try:
        for i, item in enumerate(fila, 1):
            if parar_evento.is_set():
                log_cb(f'Interrompido após {i-1} ativos. Salvando...', 'aviso')
                break

            ticker    = item['ticker']
            linha     = item['linha']
            tipo      = item['tipo']
            prefixo   = '[CART]' if item['prioridade'] else '[CAT] '

            # Estima tempo restante
            decorrido = time.time() - inicio_ts
            eta_str   = ''
            if i > 1:
                seg_por_item = decorrido / (i - 1)
                restantes    = (total - i + 1) * seg_por_item
                m, s         = divmod(int(restantes), 60)
                eta_str      = f'  ETA: {m}m {s:02d}s'

            progresso_cb(i, total, ticker, prefixo, eta_str)

            try:
                preco = obter_preco_yfinance(ticker)
                dados = obter_dados_status_invest(ticker, tipo, sessao, log_cb)

                # Fallback de DY: se Status Invest retornou 0,00, tenta yfinance
                dy_final  = dados['dy']
                dy_origem = 'SI'
                if not dy_final or dy_final in ('0,00', '0.00', '-', ''):
                    dy_yf = obter_dy_yfinance(ticker)
                    if dy_yf:
                        dy_final  = dy_yf
                        dy_origem = 'YF'

                if preco:
                    sheet.cell(row=linha, column=COL_PRECO).value = preco
                    stats['atualizados'] += 1
                    if item['prioridade']:
                        log_cb(
                            f'[CART] {ticker:<8}  R$ {preco:.2f}'
                            f'  DY: {dy_final} ({dy_origem})'
                            f'  P/VP: {dados["pvp"]}',
                            'cart'
                        )
                else:
                    stats['sem_preco'].append(ticker)

                # Short Name: busca via API oficial B3 (cotacao.b3.com.br)
                # Só busca se a célula estiver vazia — preserva valores já preenchidos
                # (sejam vindos de execução anterior ou inseridos manualmente)
                short_name_planilha = sheet.cell(row=linha, column=COL_SHORT_NAME).value
                short_name_vazio    = not short_name_planilha or not str(short_name_planilha).strip()
                if short_name_vazio:
                    short = obter_short_name_b3(ticker)
                    if short:
                        sheet.cell(row=linha, column=COL_SHORT_NAME).value = short

                # CNPJ: preenche se vazio OU atualiza se o site trouxe valor diferente
                cnpj_planilha = sheet.cell(row=linha, column=COL_CNPJ).value
                if dados['cnpj'] and dados['cnpj'] != cnpj_planilha:
                    sheet.cell(row=linha, column=COL_CNPJ).value = dados['cnpj']

                # Segmento: preenche só se vier um valor válido E a célula estiver vazia
                seg_planilha = sheet.cell(row=linha, column=COL_SEG).value
                if dados['segmento'] and len(dados['segmento']) > 2 and not seg_planilha:
                    sheet.cell(row=linha, column=COL_SEG).value = dados['segmento']

                sheet.cell(row=linha, column=COL_DY).value  = dy_final
                sheet.cell(row=linha, column=COL_PVP).value = dados['pvp']

                if i % SALVAR_A_CADA == 0:
                    wb.save(ARQUIVO_EXCEL)
                    log_cb(f'Checkpoint: {i} ativos processados, arquivo salvo.', 'info')

            except Exception as e:
                stats['erros'].append(f'{ticker}: {e}')
                continue

        wb.save(ARQUIVO_EXCEL)

    except Exception as e:
        log_cb(f'Erro inesperado: {e}', 'erro')
        try:
            wb.save(ARQUIVO_EXCEL)
        except:
            pass
        raise
    finally:
        wb.close()
        sessao.close()

    return stats


# ══════════════════════════════════════════════════════════════════════
# INTERFACE GRÁFICA
# ══════════════════════════════════════════════════════════════════════

class AtualizadorApp:
    def __init__(self, root):
        self.root        = root
        self.parar_evt   = threading.Event()
        self.em_execucao = False

        self.root.title('Atualizar Cotações — Gestão de Dividendos')
        self.root.resizable(False, False)
        self._centralizar(720, 600)
        self._construir_ui()

    # ------------------------------------------------------------------
    # Layout
    # ------------------------------------------------------------------
    def _construir_ui(self):
        # Cabeçalho
        header = tk.Frame(self.root, bg=COR_VERDE, height=60)
        header.pack(fill='x')
        header.pack_propagate(False)
        tk.Label(header, text='  Atualizar Cotações e DY',
                 bg=COR_VERDE, fg='white',
                 font=('Segoe UI', 13, 'bold'), anchor='w'
                 ).pack(fill='both', expand=True, padx=4)

        # Cards de resumo
        resumo = tk.Frame(self.root, bg='#f5f5f0', padx=20, pady=10)
        resumo.pack(fill='x')
        self.card_total      = self._card(resumo, 'Total no catálogo',    '—', 0)
        self.card_atualizados = self._card(resumo, 'Preços atualizados',  '—', 1)
        self.card_sem_preco  = self._card(resumo, 'Sem cotação',          '—', 2)
        self.card_erros      = self._card(resumo, 'Erros',                '—', 3)
        resumo.columnconfigure((0, 1, 2, 3), weight=1)

        # Progresso
        prog_frame = tk.Frame(self.root, padx=20, pady=6)
        prog_frame.pack(fill='x')

        self.var_ticker_atual = tk.StringVar(value='Aguardando início...')
        tk.Label(prog_frame, textvariable=self.var_ticker_atual,
                 font=('Segoe UI', 9, 'bold'), anchor='w'
                 ).pack(fill='x')

        self.progressbar = ttk.Progressbar(prog_frame, mode='determinate', length=680)
        self.progressbar.pack(fill='x', pady=(2, 2))

        self.var_eta = tk.StringVar(value='')
        tk.Label(prog_frame, textvariable=self.var_eta,
                 font=('Segoe UI', 8), fg='gray', anchor='w'
                 ).pack(fill='x')

        # Log
        log_frame = tk.Frame(self.root, padx=20, pady=2)
        log_frame.pack(fill='both', expand=True)

        self.log = scrolledtext.ScrolledText(
            log_frame, height=15, state='disabled',
            font=('Consolas', 9), wrap='word',
            bg='#1e1e1e', fg='#d4d4d4',
        )
        self.log.pack(fill='both', expand=True)
        self.log.tag_config('info',  foreground='#9cdcfe')
        self.log.tag_config('cart',  foreground='#4ec9b0')
        self.log.tag_config('aviso', foreground='#dcdcaa')
        self.log.tag_config('erro',  foreground='#f44747')

        # Modo de atualização
        modo_frame = tk.Frame(self.root, bg='#f5f5f0', padx=20, pady=8)
        modo_frame.pack(fill='x')

        tk.Label(modo_frame, text='Modo de atualização:',
                 bg='#f5f5f0', font=('Segoe UI', 9, 'bold')
                 ).grid(row=0, column=0, sticky='w', padx=(0, 16))

        self.var_modo = tk.StringVar(value='carteira')

        rb_cart = tk.Radiobutton(
            modo_frame, text='Apenas minha carteira  (rápido — ~1 min)',
            variable=self.var_modo, value='carteira',
            bg='#f5f5f0', font=('Segoe UI', 9), cursor='hand2',
            activebackground='#f5f5f0'
        )
        rb_cart.grid(row=0, column=1, sticky='w')

        rb_all = tk.Radiobutton(
            modo_frame, text='Catálogo completo B3  (lento — ~30 min)',
            variable=self.var_modo, value='completo',
            bg='#f5f5f0', font=('Segoe UI', 9), cursor='hand2',
            activebackground='#f5f5f0'
        )
        rb_all.grid(row=0, column=2, sticky='w', padx=(16, 0))

        # Botões
        btn_frame = tk.Frame(self.root, padx=20, pady=10)
        btn_frame.pack(fill='x')

        self.btn_iniciar = tk.Button(
            btn_frame, text='Iniciar atualização',
            command=self._iniciar,
            bg=COR_VERDE, fg='white',
            font=('Segoe UI', 10, 'bold'),
            padx=20, pady=7, cursor='hand2',
            relief='flat', activebackground=COR_VERDE_ESC
        )
        self.btn_iniciar.pack(side='left')

        self.btn_parar = tk.Button(
            btn_frame, text='Parar e salvar',
            command=self._parar,
            bg='#6b7280', fg='white',
            font=('Segoe UI', 10),
            padx=16, pady=7, cursor='hand2',
            relief='flat', state='disabled'
        )
        self.btn_parar.pack(side='left', padx=(8, 0))

        tk.Button(
            btn_frame, text='Abrir planilha',
            command=self._abrir_planilha,
            bg='#1e40af', fg='white',
            font=('Segoe UI', 10),
            padx=16, pady=7, cursor='hand2',
            relief='flat'
        ).pack(side='left', padx=(8, 0))

        self.var_status = tk.StringVar(
            value='Clique em "Iniciar atualização" para atualizar preços e DY.')
        tk.Label(btn_frame, textvariable=self.var_status,
                 font=('Segoe UI', 9), fg='gray', wraplength=340, justify='left'
                 ).pack(side='left', padx=(16, 0))

    def _card(self, parent, titulo, valor, col):
        frame = tk.Frame(parent, bg='#f5f5f0', padx=8, pady=6)
        frame.grid(row=0, column=col, sticky='nsew', padx=(0, 8))
        tk.Label(frame, text=titulo, bg='#f5f5f0',
                 font=('Segoe UI', 8), fg='gray').pack(anchor='w')
        var = tk.StringVar(value=valor)
        tk.Label(frame, textvariable=var, bg='#f5f5f0',
                 font=('Segoe UI', 11, 'bold')).pack(anchor='w')
        return var

    # ------------------------------------------------------------------
    # Log e progresso (thread-safe via root.after)
    # ------------------------------------------------------------------
    def _log(self, msg, tag='info'):
        self.root.after(0, self._escrever_log, msg, tag)

    def _escrever_log(self, msg, tag):
        hora = datetime.now().strftime('%H:%M:%S')
        self.log.config(state='normal')
        self.log.insert('end', f'[{hora}] {msg}\n', tag)
        self.log.see('end')
        self.log.config(state='disabled')

    def _progresso(self, i, total, ticker, prefixo, eta_str):
        self.root.after(0, self._atualizar_progresso, i, total, ticker, prefixo, eta_str)

    def _atualizar_progresso(self, i, total, ticker, prefixo, eta_str):
        pct = int(i / total * 100) if total else 0
        self.progressbar['value'] = pct
        self.var_ticker_atual.set(
            f'{prefixo} {ticker}   [{i}/{total}]  {pct}%')
        self.var_eta.set(eta_str)

    # ------------------------------------------------------------------
    # Controle
    # ------------------------------------------------------------------
    def _iniciar(self):
        if self.em_execucao:
            return
        self.em_execucao = True
        self.parar_evt.clear()
        self.modo_atual = self.var_modo.get()

        # Limpa estado
        self.log.config(state='normal')
        self.log.delete('1.0', 'end')
        self.log.config(state='disabled')
        self.progressbar['value'] = 0
        self.var_ticker_atual.set('Iniciando...')
        self.var_eta.set('')
        for card in (self.card_total, self.card_atualizados,
                     self.card_sem_preco, self.card_erros):
            card.set('—')

        self.btn_iniciar.config(state='disabled')
        self.btn_parar.config(state='normal')

        if self.modo_atual == 'carteira':
            self.var_status.set('Atualizando apenas os ativos da carteira...')
        else:
            self.var_status.set('Atualizando catálogo completo — isso pode levar ~30 min...')

        threading.Thread(target=self._executar, daemon=True).start()

    def _parar(self):
        self.parar_evt.set()
        self.var_status.set('Parando após o ativo atual...')
        self.btn_parar.config(state='disabled')

    def _executar(self):
        try:
            stats = executar_atualizacao(
                log_cb       = self._log,
                progresso_cb = self._progresso,
                parar_evento = self.parar_evt,
                apenas_carteira = (self.modo_atual == 'carteira'),
            )
            self.root.after(0, self._finalizar, stats)
        except Exception as e:
            self.root.after(0, self._finalizar_erro, str(e))

    def _finalizar(self, stats):
        self.em_execucao = False
        self.btn_iniciar.config(state='normal')
        self.btn_parar.config(state='disabled')
        self.progressbar['value'] = 100
        self.var_ticker_atual.set('Concluído.')
        self.var_eta.set('')

        self.card_total.set(str(stats['total']))
        self.card_atualizados.set(str(stats['atualizados']))
        self.card_sem_preco.set(str(len(stats['sem_preco'])))
        self.card_erros.set(str(len(stats['erros'])))

        # Lista ativos sem cotação no log
        if stats['sem_preco']:
            self._log(f"Sem cotação ({len(stats['sem_preco'])}): "
                      f"{', '.join(stats['sem_preco'][:15])}"
                      f"{'...' if len(stats['sem_preco']) > 15 else ''}", 'aviso')

        # Lista primeiros erros no log
        if stats['erros']:
            self._log(f"Erros ({len(stats['erros'])}):", 'erro')
            for err in stats['erros'][:5]:
                self._log(f"  {err}", 'erro')
            if len(stats['erros']) > 5:
                self._log(f"  ... e mais {len(stats['erros']) - 5} erro(s).", 'erro')

        self._log('', 'info')
        self._log('IMPORTANTE: Abra a planilha Excel e salve-a antes de usar', 'aviso')
        self._log('o Smart Aporte ou Buscar Dividendos. Isso garante que', 'aviso')
        self._log('todas as fórmulas estejam recalculadas corretamente.', 'aviso')

        self.var_status.set(
            f"Concluído. {stats['atualizados']} preços atualizados. "
            f"Abra e salve o Excel antes de usar os outros módulos.")

    def _finalizar_erro(self, mensagem):
        self.em_execucao = False
        self.btn_iniciar.config(state='normal')
        self.btn_parar.config(state='disabled')
        self.var_status.set('Erro durante a execução.')
        messagebox.showerror('Erro', mensagem)

    # ------------------------------------------------------------------
    def _abrir_planilha(self):
        """Abre a planilha no Excel para forçar o recálculo das fórmulas."""
        try:
            import subprocess, sys
            if sys.platform == 'win32':
                os.startfile(ARQUIVO_EXCEL)
            elif sys.platform == 'darwin':
                subprocess.call(['open', ARQUIVO_EXCEL])
            else:
                subprocess.call(['xdg-open', ARQUIVO_EXCEL])
        except Exception as e:
            messagebox.showerror('Erro ao abrir planilha', str(e))

    def _centralizar(self, largura, altura):
        self.root.update_idletasks()
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        self.root.geometry(
            f'{largura}x{altura}+{(sw-largura)//2}+{(sh-altura)//2}')


def main():
    root = tk.Tk()
    AtualizadorApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()
