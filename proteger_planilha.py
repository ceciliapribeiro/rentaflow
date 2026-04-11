"""
proteger_planilha.py
====================
Protege a planilha para distribuição comercial:

  - Bloqueia fórmulas e estrutura de todas as abas
  - Libera apenas as células que o usuário precisa preencher
  - Aplica senha de proteção configurável

Células LIBERADAS (editáveis pelo usuário):
  OPERAÇÕES  : A:H inteiras (dados de operações)
  DIVIDENDOS : A:J inteiras (dados de dividendos)
  APORTES    : A:C inteiras (data, valor, descrição)
  CARTEIRA   : col B (ativo ideal) e col E (qtde ideal)
  Dados B3   : totalmente bloqueada (preenchida pelos scripts)
  DASHBOARD  : totalmente bloqueada (somente leitura)

Como usar:
  python proteger_planilha.py

Será gerado um novo arquivo:  <nome>_PROTEGIDO.xlsx
O arquivo original não é modificado.
"""

import openpyxl
from openpyxl.styles import Protection
from openpyxl.worksheet.protection import SheetProtection
import shutil
import os
import sys
import getpass

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config_loader import carregar_config

# ══════════════════════════════════════════════════════════════════════
# CONFIGURAÇÃO
# ══════════════════════════════════════════════════════════════════════

_cfg           = carregar_config()
ARQUIVO_ORIGEM = _cfg['arquivo_excel']

# Senha padrão — altere antes de distribuir ou será solicitada no terminal
SENHA_PROTECAO = ''

# Células liberadas por aba  (None = aba inteira liberada)
# Formato: lista de ranges no estilo Excel  ex: ['A2:H1000', 'C5']
CELULAS_LIBERADAS = {
    'OPERAÇÕES':  ['A2:H2000'],    # todas as colunas de dados
    'DIVIDENDOS': ['A2:J5000'],    # todas as colunas de dados
    'APORTES':    ['A2:C500'],     # data, valor, descrição
    'CARTEIRA':   ['B3:B50',       # ativo ideal
                   'E3:E50'],      # qtde ideal
    'Dados B3':   [],              # totalmente bloqueada
    'DASHBOARD':  [],              # totalmente bloqueada
}

# ══════════════════════════════════════════════════════════════════════
# LÓGICA
# ══════════════════════════════════════════════════════════════════════

def proteger_planilha(senha):
    nome_base   = os.path.splitext(ARQUIVO_ORIGEM)[0]
    arquivo_out = f"{nome_base}_PROTEGIDO.xlsx"

    print(f"\nCarregando: {os.path.basename(ARQUIVO_ORIGEM)}")
    wb = openpyxl.load_workbook(ARQUIVO_ORIGEM)

    for nome_aba in wb.sheetnames:
        sheet = wb[nome_aba]
        ranges_liberados = CELULAS_LIBERADAS.get(nome_aba, [])

        print(f"  Protegendo aba: {nome_aba}...", end=' ')

        # Passo 1 — bloqueia TODAS as células da aba
        for row in sheet.iter_rows():
            for cell in row:
                cell.protection = Protection(locked=True)

        # Passo 2 — desbloqueia os ranges editáveis
        for cell_range in ranges_liberados:
            for row_tuple in sheet[cell_range]:
                # sheet[A1:B10] sempre retorna tuple de tuples
                if isinstance(row_tuple, tuple):
                    for cell in row_tuple:
                        cell.protection = Protection(locked=False)
                else:
                    # célula única (ex: sheet['A1'])
                    row_tuple.protection = Protection(locked=False)

        # Passo 3 — ativa proteção da aba
        sheet.protection = SheetProtection(
            password           = senha,
            sheet              = True,
            objects            = True,
            scenarios          = True,
            formatCells        = False,   # permite formatar células
            formatColumns      = False,   # permite ajustar colunas
            formatRows         = False,   # permite ajustar linhas
            insertColumns      = False,
            insertRows         = True,    # permite inserir linhas (necessário para novos dados)
            deleteRows         = True,    # permite apagar linhas de dados
            sort               = False,
            autoFilter         = False,
            selectLockedCells  = True,    # permite selecionar células bloqueadas
            selectUnlockedCells= True,    # permite selecionar células liberadas
        )

        liberadas_str = ', '.join(ranges_liberados) if ranges_liberados else 'nenhuma (somente leitura)'
        print(f"OK  |  Células editáveis: {liberadas_str}")

    print(f"\nSalvando: {os.path.basename(arquivo_out)}")
    wb.save(arquivo_out)
    wb.close()

    tamanho = os.path.getsize(arquivo_out) / 1024
    print(f"\n✅ Concluído!")
    print(f"   Arquivo gerado : {arquivo_out}")
    print(f"   Tamanho        : {tamanho:,.0f} KB")
    print(f"   Senha aplicada : {'(configurada)' if senha else '(sem senha — defina SENHA_PROTECAO no script)'}")
    print()
    print("IMPORTANTE: Teste o arquivo antes de distribuir.")
    print("Verifique que os scripts conseguem gravar nas abas protegidas.")
    print("(Os scripts usam openpyxl sem senha — proteção é só para o Excel.)")


def main():
    global SENHA_PROTECAO

    print("=" * 52)
    print("  PROTEÇÃO DE PLANILHA — GESTÃO DE DIVIDENDOS")
    print("=" * 52)

    if not os.path.isfile(ARQUIVO_ORIGEM):
        print(f"\n[ERRO] Arquivo não encontrado:\n  {ARQUIVO_ORIGEM}")
        print("Verifique o config.json e tente novamente.")
        input("\nPressione Enter para sair...")
        return

    # Solicita senha se não configurada
    if not SENHA_PROTECAO:
        print("\nDefina a senha de proteção da planilha.")
        print("(O usuário precisará desta senha para editar fórmulas e estrutura)\n")
        while True:
            senha1 = getpass.getpass("Senha: ")
            senha2 = getpass.getpass("Confirme: ")
            if not senha1:
                print("[AVISO] Senha vazia — a planilha ficará sem proteção por senha.")
                SENHA_PROTECAO = ''
                break
            if senha1 == senha2:
                SENHA_PROTECAO = senha1
                break
            print("As senhas não coincidem. Tente novamente.\n")

    proteger_planilha(SENHA_PROTECAO)
    input("Pressione Enter para sair...")


if __name__ == '__main__':
    main()
