"""
Prepara a planilha para distribuição como produto RentaFlow:
  - Protege as fórmulas de todas as abas
  - Libera apenas as células de dados que o cliente precisa editar
  - Sem senha por enquanto (adicione SENHA_PROTECAO quando definir)
"""
import openpyxl
from openpyxl.styles import Protection, Font, PatternFill, Alignment
from openpyxl.worksheet.protection import SheetProtection
import shutil, os

ORIGEM  = '/home/claude/RentaFlow_Planilha.xlsx'
DESTINO = '/home/claude/RentaFlow_Planilha_CLIENTE.xlsx'
SENHA   = ''rainhadasplanilhas   # ← adicione a senha aqui quando definir

# Ranges editáveis por aba (fora desses, tudo fica bloqueado)
EDITAVEL = {
    'Dados B3':   [],                        # somente leitura — atualizado pelo app
    'CARTEIRA':   ['B3:B50', 'E3:E50'],      # ATIVO ideal e QTDE ideal
    'OPERAÇÕES':  ['A2:K2000'],              # dados de operações
    'DIVIDENDOS': ['A2:G5000'],             # dados de dividendos
    'APORTES':    ['A2:C500'],              # data, valor, descrição
    'DASHBOARD':  [],                        # somente leitura
}

shutil.copy2(ORIGEM, DESTINO)
wb = openpyxl.load_workbook(DESTINO)

for nome_aba in wb.sheetnames:
    sheet = wb[nome_aba]
    ranges = EDITAVEL.get(nome_aba, [])

    # 1. Bloqueia todas as células
    for row in sheet.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)

    # 2. Desbloqueia os ranges editáveis
    for rng in ranges:
        for row_tuple in sheet[rng]:
            cells = row_tuple if isinstance(row_tuple, tuple) else [row_tuple]
            for cell in cells:
                cell.protection = Protection(locked=False)

    # 3. Ativa proteção da aba
    sheet.protection = SheetProtection(
        password           = SENHA,
        sheet              = True,
        objects            = True,
        scenarios          = True,
        formatCells        = False,
        formatColumns      = False,
        formatRows         = False,
        insertRows         = True,
        deleteRows         = True,
        selectLockedCells  = True,
        selectUnlockedCells= True,
    )

    modo = ', '.join(ranges) if ranges else 'somente leitura'
    print(f'  ✅ {nome_aba:<14} protegida  |  editável: {modo}')

wb.save(DESTINO)
wb.close()
print(f'\nSalvo em: {DESTINO}')
print(f'Tamanho: {os.path.getsize(DESTINO)/1024:.0f} KB')
