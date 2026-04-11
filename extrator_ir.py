import openpyxl
import pandas as pd
import os
import sys
import shutil
import threading
import subprocess
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config_loader import carregar_config, fazer_backup

_cfg           = carregar_config()
ARQUIVO_ORIGEM = _cfg['arquivo_excel']

# ==========================================
# CONFIGURAÇÃO DE MAPEAMENTO
# ==========================================
ABA_OPERACOES  = 'OPERAÇÕES'
ABA_DIVIDENDOS = 'DIVIDENDOS'
ABA_DADOS_FII  = 'Dados B3'

COL_DATA_OP    = 1
COL_TICKER_OP  = 2
COL_QTDE_OP    = 3
COL_PRECO_UNIT = 4
COL_TIPO_ATIVO = 5
COL_OPERACAO   = 8

COL_ANO_DIV    = 1
COL_VALOR_DIV  = 5
COL_TIPO_DIV   = 6
COL_ATIVO_DIV  = 7

COL_TICKER_REF = 1
COL_RAZAO_REF  = 3
COL_CNPJ_REF   = 5

COR_VERDE      = '#1a6b45'
COR_VERDE_ESC  = '#145535'
# ==========================================


# ══════════════════════════════════════════════════════════════════════
# LÓGICA DE NEGÓCIO
# ══════════════════════════════════════════════════════════════════════

def gerar_relatorio_ir(ano_base, log_cb):
    """
    Processa os dados e retorna (bens_texto, rend_list, avisos, caminho_arquivo).
    Não salva nada — a UI decide quando salvar após o preview.
    """
    data_limite = datetime(ano_base, 12, 31)
    avisos      = []

    log_cb(f'Carregando planilha para o ano {ano_base}...', 'info')

    try:
        wb = openpyxl.load_workbook(ARQUIVO_ORIGEM, data_only=True)
    except Exception as e:
        raise RuntimeError(f'Erro ao abrir a planilha: {e}')

    # 1 — Metadados (razão social e CNPJ)
    log_cb('Indexando razão social e CNPJ...', 'info')
    sheet_ref    = wb[ABA_DADOS_FII]
    index_ativos = {}
    for row in range(2, sheet_ref.max_row + 1):
        t = str(sheet_ref.cell(row=row, column=COL_TICKER_REF).value).strip().upper()
        if t and t != 'NONE':
            index_ativos[t] = {
                'razao': sheet_ref.cell(row=row, column=COL_RAZAO_REF).value,
                'cnpj':  sheet_ref.cell(row=row, column=COL_CNPJ_REF).value,
            }

    # 2 — Bens e Direitos: reconstrói saldo até 31/12/ano_base
    log_cb(f'Reconstruindo custódia até 31/12/{ano_base}...', 'info')
    sheet_ops     = wb[ABA_OPERACOES]
    carteira_hist = {}
    erros_op      = 0

    for row in range(2, sheet_ops.max_row + 1):
        data_raw = sheet_ops.cell(row=row, column=COL_DATA_OP).value
        if isinstance(data_raw, str):
            try:
                data_op = datetime.strptime(data_raw, '%Y-%m-%d')
            except:
                continue
        else:
            data_op = data_raw

        if not data_op or data_op > data_limite:
            continue

        ticker = str(sheet_ops.cell(row=row, column=COL_TICKER_OP).value).strip().upper()
        try:
            qtde    = float(sheet_ops.cell(row=row, column=COL_QTDE_OP).value or 0)
            preco   = float(sheet_ops.cell(row=row, column=COL_PRECO_UNIT).value or 0)
            op_tipo = str(sheet_ops.cell(row=row, column=COL_OPERACAO).value).upper()
            classe  = str(sheet_ops.cell(row=row, column=COL_TIPO_ATIVO).value).upper()
        except:
            continue

        if ticker not in carteira_hist:
            carteira_hist[ticker] = {'qtde': 0, 'custo': 0.0, 'tipo': classe}

        if 'COMPRA' in op_tipo:
            carteira_hist[ticker]['qtde']  += qtde
            carteira_hist[ticker]['custo'] += qtde * preco
        elif 'VENDA' in op_tipo:
            qtde_atual = carteira_hist[ticker]['qtde']
            if qtde_atual <= 0:
                aviso = f'Linha {row}: VENDA de {ticker} sem compra anterior — ignorada.'
                log_cb(aviso, 'aviso')
                avisos.append(aviso)
                erros_op += 1
                continue
            c_medio = carteira_hist[ticker]['custo'] / qtde_atual
            carteira_hist[ticker]['qtde']  -= qtde
            carteira_hist[ticker]['custo'] -= qtde * c_medio

    if erros_op:
        log_cb(f'{erros_op} operação(ões) com inconsistência ignorada(s).', 'aviso')

    bens_texto = []
    for ticker, dados in sorted(carteira_hist.items()):
        if dados['qtde'] > 0.01:
            meta  = index_ativos.get(ticker, {'razao': ticker, 'cnpj': 'Consultar RI'})
            razao = meta['razao'] or ticker
            cnpj  = meta['cnpj']  or 'Consultar RI'
            bens_texto.append({
                'ticker': ticker,
                'tipo':   dados['tipo'],
                'qtde':   int(dados['qtde']),
                'custo':  round(dados['custo'], 2),
                'razao':  razao,
                'cnpj':   cnpj,
                'discriminacao': (
                    f"{int(dados['qtde'])} COTAS DO {dados['tipo']} {ticker} - "
                    f"{razao}, CNPJ: {cnpj}. "
                    f"CUSTO TOTAL DE AQUISIÇÃO: R$ {dados['custo']:.2f}"
                ),
            })

    log_cb(f'{len(bens_texto)} ativo(s) em custódia em 31/12/{ano_base}.', 'info')

    # 3 — Rendimentos do ano
    log_cb(f'Consolidando rendimentos de {ano_base}...', 'info')
    sheet_div        = wb[ABA_DIVIDENDOS]
    rend_consolidado = {}

    for row in range(2, sheet_div.max_row + 1):
        try:
            ano_celula = int(sheet_div.cell(row=row, column=COL_ANO_DIV).value)
            if ano_celula != ano_base:
                continue
            ativo  = str(sheet_div.cell(row=row, column=COL_ATIVO_DIV).value).upper().strip()
            valor  = float(sheet_div.cell(row=row, column=COL_VALOR_DIV).value or 0)
            tipo_p = str(sheet_div.cell(row=row, column=COL_TIPO_DIV).value).upper()
            if ativo not in rend_consolidado:
                rend_consolidado[ativo] = {'isentos': 0.0, 'jcp': 0.0}
            if 'JUROS' in tipo_p or 'JCP' in tipo_p:
                rend_consolidado[ativo]['jcp']     += valor
            else:
                rend_consolidado[ativo]['isentos'] += valor
        except:
            continue

    rend_list = [
        {'ativo': k, 'isentos': v['isentos'], 'jcp': v['jcp'],
         'total': round(v['isentos'] + v['jcp'], 2)}
        for k, v in sorted(rend_consolidado.items())
    ]

    total_isentos = sum(r['isentos'] for r in rend_list)
    total_jcp     = sum(r['jcp']     for r in rend_list)
    log_cb(f'Rendimentos isentos: R$ {total_isentos:,.2f}  |  JCP: R$ {total_jcp:,.2f}', 'info')

    wb.close()

    nome_base       = os.path.splitext(ARQUIVO_ORIGEM)[0]
    caminho_arquivo = f"{nome_base}_IR_{ano_base}.xlsx"

    return bens_texto, rend_list, avisos, caminho_arquivo


def salvar_relatorio(bens_texto, rend_list, ano_base, caminho_arquivo, log_cb):
    fazer_backup(ARQUIVO_ORIGEM, log_cb)

    df_bens = pd.DataFrame(
        [b['discriminacao'] for b in bens_texto],
        columns=[f'Discriminação em 31/12/{ano_base}']
    )
    df_rend = pd.DataFrame([{
        'Ativo':                         r['ativo'],
        'Rendimentos Isentos':           r['isentos'],
        'JCP (Tributação Exclusiva)':    r['jcp'],
    } for r in rend_list])

    with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
        df_bens.to_excel(writer, sheet_name='Bens e Direitos',   index=False)
        df_rend.to_excel(writer, sheet_name='Rendimentos Anuais', index=False)

    log_cb(f'Relatório salvo: {os.path.basename(caminho_arquivo)}', 'info')


# ══════════════════════════════════════════════════════════════════════
# INTERFACE GRÁFICA
# ══════════════════════════════════════════════════════════════════════

class ExtratorIRApp:
    def __init__(self, root):
        self.root            = root
        self.bens_gerados    = []
        self.rend_gerados    = []
        self.caminho_arquivo = ''
        self.ano_base        = datetime.now().year - 1

        self.root.title('Relatório de IR — Imposto de Renda')
        self.root.resizable(False, False)
        self._centralizar(780, 640)
        self._construir_ui()

    # ------------------------------------------------------------------
    # Layout
    # ------------------------------------------------------------------
    def _construir_ui(self):
        # Cabeçalho
        header = tk.Frame(self.root, bg=COR_VERDE, height=60)
        header.pack(fill='x')
        header.pack_propagate(False)
        tk.Label(header, text='  Relatório de Imposto de Renda',
                 bg=COR_VERDE, fg='white',
                 font=('Segoe UI', 13, 'bold'), anchor='w'
                 ).pack(fill='both', expand=True, padx=4)

        # Seletor de ano + cards
        topo = tk.Frame(self.root, padx=20, pady=12)
        topo.pack(fill='x')

        tk.Label(topo, text='Ano-calendário:',
                 font=('Segoe UI', 10)).grid(row=0, column=0, sticky='w')

        self.var_ano = tk.IntVar(value=self.ano_base)
        anos         = list(range(datetime.now().year - 1, datetime.now().year - 6, -1))
        cb_ano = ttk.Combobox(topo, textvariable=self.var_ano,
                              values=anos, width=8, state='readonly',
                              font=('Segoe UI', 10))
        cb_ano.grid(row=0, column=1, padx=(8, 40), sticky='w')

        # Cards de resumo
        self.card_ativos    = self._card(topo, 'Ativos em custódia', '—', 2)
        self.card_isentos   = self._card(topo, 'Rendimentos isentos', '—', 3)
        self.card_jcp       = self._card(topo, 'JCP (tributável)', '—', 4)
        topo.columnconfigure((2, 3, 4), weight=1)

        # Notebook: Bens e Direitos | Rendimentos | Log
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True, padx=20, pady=(0, 4))

        # Aba 1 — Bens e Direitos
        aba_bens = tk.Frame(self.notebook)
        self.notebook.add(aba_bens, text='  Bens e Direitos  ')

        cols_b = ('ticker', 'tipo', 'qtde', 'custo', 'razao', 'cnpj')
        self.tree_bens = ttk.Treeview(aba_bens, columns=cols_b,
                                      show='headings', height=12)
        defs_b = [
            ('ticker', 'Ticker',        70,  'center'),
            ('tipo',   'Tipo',          60,  'center'),
            ('qtde',   'Qtde',          55,  'center'),
            ('custo',  'Custo total',   100, 'e'),
            ('razao',  'Razão Social',  250, 'w'),
            ('cnpj',   'CNPJ',          130, 'center'),
        ]
        for col, titulo, larg, ancora in defs_b:
            self.tree_bens.heading(col, text=titulo)
            self.tree_bens.column(col, width=larg, anchor=ancora, stretch=False)
        self.tree_bens.tag_configure('par',   background='#f9f9f6')
        self.tree_bens.tag_configure('impar', background='#ffffff')
        scroll_b = ttk.Scrollbar(aba_bens, orient='vertical',
                                 command=self.tree_bens.yview)
        self.tree_bens.configure(yscrollcommand=scroll_b.set)
        self.tree_bens.pack(side='left', fill='both', expand=True)
        scroll_b.pack(side='right', fill='y')

        # Aba 2 — Rendimentos Anuais
        aba_rend = tk.Frame(self.notebook)
        self.notebook.add(aba_rend, text='  Rendimentos Anuais  ')

        cols_r = ('ativo', 'isentos', 'jcp', 'total')
        self.tree_rend = ttk.Treeview(aba_rend, columns=cols_r,
                                      show='headings', height=12)
        defs_r = [
            ('ativo',   'Ativo',               100, 'center'),
            ('isentos', 'Rendimentos Isentos', 180, 'e'),
            ('jcp',     'JCP (Tributável)',     180, 'e'),
            ('total',   'Total recebido',       150, 'e'),
        ]
        for col, titulo, larg, ancora in defs_r:
            self.tree_rend.heading(col, text=titulo)
            self.tree_rend.column(col, width=larg, anchor=ancora, stretch=False)
        self.tree_rend.tag_configure('par',   background='#f9f9f6')
        self.tree_rend.tag_configure('impar', background='#ffffff')
        scroll_r = ttk.Scrollbar(aba_rend, orient='vertical',
                                 command=self.tree_rend.yview)
        self.tree_rend.configure(yscrollcommand=scroll_r.set)
        self.tree_rend.pack(side='left', fill='both', expand=True)
        scroll_r.pack(side='right', fill='y')

        # Aba 3 — Log
        aba_log = tk.Frame(self.notebook)
        self.notebook.add(aba_log, text='  Log  ')
        self.log = tk.Text(aba_log, height=12, state='disabled',
                           font=('Consolas', 9), wrap='word',
                           bg='#1e1e1e', fg='#d4d4d4')
        scroll_l = ttk.Scrollbar(aba_log, orient='vertical',
                                 command=self.log.yview)
        self.log.configure(yscrollcommand=scroll_l.set)
        self.log.pack(side='left', fill='both', expand=True)
        scroll_l.pack(side='right', fill='y')
        self.log.tag_config('info',  foreground='#9cdcfe')
        self.log.tag_config('aviso', foreground='#dcdcaa')
        self.log.tag_config('erro',  foreground='#f44747')

        # Rodapé
        btn_frame = tk.Frame(self.root, padx=20, pady=10)
        btn_frame.pack(fill='x')

        self.btn_gerar = tk.Button(
            btn_frame, text='1. Gerar preview',
            command=self._gerar,
            bg='#374151', fg='white',
            font=('Segoe UI', 10, 'bold'),
            padx=18, pady=7, cursor='hand2', relief='flat'
        )
        self.btn_gerar.pack(side='left')

        self.btn_salvar = tk.Button(
            btn_frame, text='2. Salvar relatório (.xlsx)',
            command=self._salvar,
            bg=COR_VERDE, fg='white',
            font=('Segoe UI', 10, 'bold'),
            padx=18, pady=7, cursor='hand2',
            relief='flat', activebackground=COR_VERDE_ESC,
            state='disabled'
        )
        self.btn_salvar.pack(side='left', padx=(8, 0))

        self.btn_abrir = tk.Button(
            btn_frame, text='Abrir arquivo',
            command=self._abrir_arquivo,
            bg='#1e40af', fg='white',
            font=('Segoe UI', 10),
            padx=16, pady=7, cursor='hand2',
            relief='flat', state='disabled'
        )
        self.btn_abrir.pack(side='left', padx=(8, 0))

        self.var_status = tk.StringVar(
            value='Selecione o ano e clique em "1. Gerar preview".')
        tk.Label(btn_frame, textvariable=self.var_status,
                 font=('Segoe UI', 9), fg='gray', wraplength=280
                 ).pack(side='left', padx=(16, 0))

    def _card(self, parent, titulo, valor, col):
        frame = tk.Frame(parent, bg='#f5f5f0', padx=8, pady=4)
        frame.grid(row=0, column=col, sticky='nsew', padx=(0, 8))
        tk.Label(frame, text=titulo, bg='#f5f5f0',
                 font=('Segoe UI', 8), fg='gray').pack(anchor='w')
        var = tk.StringVar(value=valor)
        tk.Label(frame, textvariable=var, bg='#f5f5f0',
                 font=('Segoe UI', 11, 'bold')).pack(anchor='w')
        return var

    # ------------------------------------------------------------------
    # Log
    # ------------------------------------------------------------------
    def _log(self, msg, tag='info'):
        self.root.after(0, self._escrever_log, msg, tag)

    def _escrever_log(self, msg, tag):
        hora = datetime.now().strftime('%H:%M:%S')
        self.log.config(state='normal')
        self.log.insert('end', f'[{hora}] {msg}\n', tag)
        self.log.see('end')
        self.log.config(state='disabled')

    # ------------------------------------------------------------------
    # Passo 1 — Preview
    # ------------------------------------------------------------------
    def _gerar(self):
        self.ano_base = self.var_ano.get()

        # Limpa estado
        for tree in (self.tree_bens, self.tree_rend):
            for row in tree.get_children():
                tree.delete(row)
        self.log.config(state='normal')
        self.log.delete('1.0', 'end')
        self.log.config(state='disabled')
        for card in (self.card_ativos, self.card_isentos, self.card_jcp):
            card.set('—')
        self.btn_salvar.config(state='disabled')
        self.btn_abrir.config(state='disabled')
        self.btn_gerar.config(state='disabled', text='Processando...')
        self.var_status.set('Processando...')
        self.notebook.select(2)  # vai para aba Log

        threading.Thread(target=self._processar, daemon=True).start()

    def _processar(self):
        try:
            bens, rend, avisos, caminho = gerar_relatorio_ir(
                self.ano_base, self._log)
            self.root.after(0, self._exibir_preview, bens, rend, avisos, caminho)
        except Exception as e:
            self.root.after(0, self._finalizar_erro, str(e))

    def _exibir_preview(self, bens, rend, avisos, caminho):
        self.bens_gerados    = bens
        self.rend_gerados    = rend
        self.caminho_arquivo = caminho

        # Popula Bens e Direitos
        for i, b in enumerate(bens):
            tag = 'par' if i % 2 == 0 else 'impar'
            self.tree_bens.insert('', 'end', tags=(tag,), values=(
                b['ticker'], b['tipo'], b['qtde'],
                f"R$ {b['custo']:,.2f}", b['razao'], b['cnpj'],
            ))

        # Popula Rendimentos
        for i, r in enumerate(rend):
            tag = 'par' if i % 2 == 0 else 'impar'
            self.tree_rend.insert('', 'end', tags=(tag,), values=(
                r['ativo'],
                f"R$ {r['isentos']:,.2f}",
                f"R$ {r['jcp']:,.2f}",
                f"R$ {r['total']:,.2f}",
            ))

        # Atualiza cards
        total_isentos = sum(r['isentos'] for r in rend)
        total_jcp     = sum(r['jcp']     for r in rend)
        self.card_ativos.set(str(len(bens)))
        self.card_isentos.set(f'R$ {total_isentos:,.2f}')
        self.card_jcp.set(f'R$ {total_jcp:,.2f}')

        self.btn_gerar.config(state='normal', text='1. Gerar preview')
        self.btn_salvar.config(state='normal')
        self.notebook.select(0)  # vai para aba Bens e Direitos

        aviso_txt = f'  |  {len(avisos)} aviso(s)' if avisos else ''
        self.var_status.set(
            f'Preview pronto. Revise e clique em "2. Salvar relatório"{aviso_txt}.')

    # ------------------------------------------------------------------
    # Passo 2 — Salvar
    # ------------------------------------------------------------------
    def _salvar(self):
        if not self.bens_gerados and not self.rend_gerados:
            return

        confirmado = messagebox.askyesno(
            'Confirmar geração',
            f'Gerar o relatório de IR para o ano {self.ano_base}?\n\n'
            f'Arquivo: {os.path.basename(self.caminho_arquivo)}\n'
            f'Local: {os.path.dirname(self.caminho_arquivo)}'
        )
        if not confirmado:
            return

        self.btn_salvar.config(state='disabled', text='Salvando...')
        self.notebook.select(2)

        threading.Thread(target=self._gravar, daemon=True).start()

    def _gravar(self):
        try:
            salvar_relatorio(
                self.bens_gerados, self.rend_gerados,
                self.ano_base, self.caminho_arquivo, self._log
            )
            self.root.after(0, self._finalizar)
        except Exception as e:
            self.root.after(0, self._finalizar_erro, str(e))

    def _finalizar(self):
        self.btn_salvar.config(state='normal', text='2. Salvar relatório (.xlsx)')
        self.btn_abrir.config(state='normal')
        self.var_status.set(
            f'Relatório gerado: {os.path.basename(self.caminho_arquivo)}')
        messagebox.showinfo(
            'Relatório gerado',
            f'Arquivo salvo com sucesso:\n{self.caminho_arquivo}'
        )

    def _finalizar_erro(self, mensagem):
        self.btn_gerar.config(state='normal', text='1. Gerar preview')
        self.btn_salvar.config(state='normal', text='2. Salvar relatório (.xlsx)')
        self.var_status.set('Erro durante o processamento.')
        messagebox.showerror('Erro', mensagem)

    # ------------------------------------------------------------------
    # Abrir arquivo gerado
    # ------------------------------------------------------------------
    def _abrir_arquivo(self):
        if not self.caminho_arquivo or not os.path.exists(self.caminho_arquivo):
            messagebox.showwarning('Arquivo não encontrado',
                                   'O arquivo ainda não foi gerado ou foi movido.')
            return
        try:
            if sys.platform == 'win32':
                os.startfile(self.caminho_arquivo)
            elif sys.platform == 'darwin':
                subprocess.call(['open', self.caminho_arquivo])
            else:
                subprocess.call(['xdg-open', self.caminho_arquivo])
        except Exception as e:
            messagebox.showerror('Erro ao abrir', str(e))

    # ------------------------------------------------------------------
    def _centralizar(self, largura, altura):
        self.root.update_idletasks()
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        self.root.geometry(
            f'{largura}x{altura}+{(sw-largura)//2}+{(sh-altura)//2}')


def main():
    root = tk.Tk()
    ExtratorIRApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()
